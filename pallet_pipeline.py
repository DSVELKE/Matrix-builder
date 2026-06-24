"""
pallet_pipeline.py — build DHL freight (pallet / EUROCONNECT) matrices.
 
Kept separate from pipeline.py so the parcel logic is untouched. Reuses
pipeline.py's Excel writer column conventions but adds the pallet-only cost
columns (MOBILITY, TOLL, ADMIN, FACTORED RATE PALLET).
 
Cost stack (verified to the cent against ITFRDE_final_for_Fender_Pallets.xlsx):
 
    RATE_BASE = factored rate  (from the DHL factor file, already * FACTOR_DHL)
    FUEL      = FUEL_PCT      * RATE_BASE
    MOBILITY  = MOBILITY_PCT  * RATE_BASE
    MAUT      = maut_pct(country, weight) * RATE_BASE   # 2-tier (e.g. DE @2500kg)
    TOLL      = TOLL_PCT(country)  * RATE_BASE           # UK only in the contract
    ADMIN     = ADMIN_FLAT (€ per row)
    TOTAL     = RATE_BASE + FUEL + MOBILITY + MAUT + TOLL + ADMIN
"""
 
import logging
from copy import deepcopy
from pathlib import Path
 
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openpyxl
 
import pallet_parser as pp
 
log = logging.getLogger(__name__)
 
 
# ==============================================================================
# CONFIG  (all of this comes from the matrix 'Variables' sheet)
# ==============================================================================
 
PALLET_GLOBALS = {
    'CARRIER_ID':   'DHL-FENDER',
    'SERVICE_LEVEL': 'EUROCONNECT',
    'SITE_ID':      'NLMOE01',
    'CLIENT_ID':    'NLFENDER',
    'FUEL_PCT':     0.155,    # FUEL DHL
    'MOBILITY_PCT': 0.04,     # MOBILITY PALLET
    'ADMIN_FLAT':   46.51,    # ADMIN PALLET (€ per row)
    'FACTOR_DHL':   1.0,      # rates in the factor file are already factored
    'TOLL_UK_PCT':  0.0043,   # TOLL UK PALLET (applied to GB only)
}
 
# Per-country MAUT as a % of RATE_BASE, with an optional 2nd tier above a weight
# breakpoint.  KNOWN values are taken straight from the reference matrix's
# Variables sheet.  Countries marked _UNKNOWN must be filled before trusting the
# MAUT column — they currently default to 0 and raise a warning.
#   (low_pct, high_pct, tier_kg)   high_pct applies to bands whose ceiling > tier_kg
MAUT_PCT = {
    'DE': (0.0253, 0.0544, 2500),
    'IT': (0.0253, 0.0253, 2500),
    'FR': (0.0,    0.0,     2500),
}
# Countries that get the UK toll
TOLL_COUNTRIES = {'GB'}
 
# Column layout for pallet output (superset of the parcel columns, in the order
# the reference file uses).
PALLET_COLUMNS = [
    'SITE_ID', 'CLIENT_ID', 'CARRIER_ID', 'SERVICE_LEVEL', 'COUNTRYISO2',
    'POSTCODE', 'MIN_WEIGHT', 'MAX_WEIGHT', 'MIN_VOLUME', 'MAX_VOLUME',
    'MIN_PARCEL', 'MAX_PARCEL', 'EACH_WEIGHT', 'EACH_VOLUME',
    'FACTORED RATE PALLET', 'USER_DEF_TYPE_1', 'USER_DEF_TYPE_2',
    'USER_DEF_TYPE_4 (max 1,5m)', 'AWKWARD',
    'RATE_BASE', 'RATE_EXTRA', 'MOBILITY', 'FUEL', 'MAUT', 'Linehaul UPSDE',
    'TOLL', 'ADMIN', 'TOTAL_PRICE',
]
 
PALLET_VARIABLES_LAYOUT = [
    ('FUEL DHL',        PALLET_GLOBALS['FUEL_PCT']),
    ('MOBILITY PALLET', PALLET_GLOBALS['MOBILITY_PCT']),
    ('ADMIN PALLET',    PALLET_GLOBALS['ADMIN_FLAT']),
    ('FACTOR DHL',      PALLET_GLOBALS['FACTOR_DHL']),
    ('TOLL UK PALLET',  PALLET_GLOBALS['TOLL_UK_PCT']),
    (None, None),
    ('MAUT DHL (low / high tiers per country are stamped numerically)', None),
]
 
 
def _maut_for(country, ceiling_kg, maut_table):
    rule = maut_table.get(country.upper())
    if rule is None:
        return None            # signals "unknown — warn"
    low, high, tier = rule
    return high if (ceiling_kg is not None and ceiling_kg > tier) else low
 
 
# ==============================================================================
# BUILDER
# ==============================================================================
 
def build_rows_pallet(country, zip_rate_map, bands, globals_=None, maut_table=None):
    """One row per (zip-prefix, weight band) straight from the factor file.
 
    zip_rate_map : {zip_prefix(str): {ceiling_kg(int): factored_rate}}
    bands        : ordered list of band ceilings (kg)
    """
    g = globals_ or PALLET_GLOBALS
    mt = maut_table or MAUT_PCT
    iso = country.upper()
    rows = []
    maut_known = iso in mt
 
    for zkey in sorted(zip_rate_map, key=lambda z: (len(z), z)):
        band_map = zip_rate_map[zkey]
        prev_ceiling = 0
        for ceil in bands:
            rate = band_map.get(ceil)
            if rate is None:
                continue
            rate_base = round(rate * g['FACTOR_DHL'], 6)
            fuel = round(g['FUEL_PCT'] * rate_base, 6)
            mob = round(g['MOBILITY_PCT'] * rate_base, 6)
            maut_pct = _maut_for(iso, ceil, mt)
            maut = round((maut_pct or 0.0) * rate_base, 8)
            toll_pct = g['TOLL_UK_PCT'] if iso in TOLL_COUNTRIES else 0.0
            toll = round(toll_pct * rate_base, 6)
            admin = g['ADMIN_FLAT']
            total = round(rate_base + fuel + mob + maut + toll + admin, 6)
            rows.append({
                'SITE_ID': g['SITE_ID'], 'CLIENT_ID': g['CLIENT_ID'],
                'CARRIER_ID': g['CARRIER_ID'], 'SERVICE_LEVEL': g['SERVICE_LEVEL'],
                'COUNTRYISO2': iso, 'POSTCODE': zkey,
                'MIN_WEIGHT': prev_ceiling if prev_ceiling > 0 else None,
                'MAX_WEIGHT': ceil,
                'MIN_VOLUME': None, 'MAX_VOLUME': None,
                'MIN_PARCEL': None, 'MAX_PARCEL': None,
                'EACH_WEIGHT': None, 'EACH_VOLUME': None,
                'FACTORED RATE PALLET': rate_base,
                'USER_DEF_TYPE_1': None, 'USER_DEF_TYPE_2': None,
                'USER_DEF_TYPE_4 (max 1,5m)': None, 'AWKWARD': None,
                'RATE_BASE': rate_base, 'RATE_EXTRA': 0,
                'MOBILITY': mob, 'FUEL': fuel, 'MAUT': maut,
                'Linehaul UPSDE': None, 'TOLL': toll, 'ADMIN': admin,
                'TOTAL_PRICE': total,
            })
            prev_ceiling = ceil
    return rows, maut_known
 
 
# ==============================================================================
# OPTIMIZER  (collapse adjacent same-rate weight bands per zip)
# ==============================================================================
 
def optimize_pallet(df):
    """Within each (country, zip), drop a band whose RATE_BASE equals the band
    directly below it (the cheaper/lighter one already covers those shipments
    because CargoWrite matches the first row whose MAX_WEIGHT >= shipment)."""
    if df.empty:
        return df
    df = df.sort_values(['COUNTRYISO2', 'POSTCODE', 'MAX_WEIGHT'],
                        kind='stable').reset_index(drop=True)
    keep = []
    last_key = None
    last_rate = None
    for i, r in df.iterrows():
        key = (r['COUNTRYISO2'], r['POSTCODE'])
        if key == last_key and r['RATE_BASE'] == last_rate:
            continue            # same price as the lighter band -> redundant
        keep.append(i)
        last_key, last_rate = key, r['RATE_BASE']
    return df.loc[keep].reset_index(drop=True)
 
 
# ==============================================================================
# EXCEL WRITER
# ==============================================================================
 
def write_pallet_excel(df, output_path, country, variables_layout=None):
    vl = variables_layout or PALLET_VARIABLES_LAYOUT
    wb = Workbook()
    ws = wb.active
    ws.title = f"{country.upper()} Pallet"
    for ci, col in enumerate(PALLET_COLUMNS, 1):
        ws.cell(1, ci, col)
    df_sorted = df.sort_values('TOTAL_PRICE', kind='stable').reset_index(drop=True)
    bucket_fill = PatternFill('solid', fgColor='FFF2CC')
    for ri, rec in enumerate(df_sorted.to_dict('records'), start=2):
        is_bucket = bool(rec.get('_is_bucket'))
        for ci, col in enumerate(PALLET_COLUMNS, 1):
            val = rec.get(col)
            cell = ws.cell(ri, ci, None if (val is None or (isinstance(val, float) and pd.isna(val))) else val)
            if is_bucket:
                cell.fill = bucket_fill
    vs = wb.create_sheet('Variables')
    for ri, (name, val) in enumerate(vl, 1):
        vs.cell(ri, 1, name)
        if val is not None:
            vs.cell(ri, 2, val)
    # also write the per-country MAUT rules used, for transparency
    vs.cell(len(vl) + 2, 1, 'MAUT rules used (low, high, tier_kg):')
    for j, (cc, rule) in enumerate(MAUT_PCT.items(), start=len(vl) + 3):
        vs.cell(j, 1, cc)
        vs.cell(j, 2, str(rule))
    wb.save(output_path)
    log.info('wrote %s (%d rows)', output_path, len(df_sorted))
 
 
# ==============================================================================
# ORCHESTRATOR
# ==============================================================================
 
def run_pallet_pipeline(parsed, country, output_dir='.',
                        globals_=None, maut_table=None, optimize=True):
    """Build extended + optimized pallet matrices for one country.
 
    parsed : output of pallet_parser.parse_pallet_factor_file()
    Returns dict: extended, optimized, rows_extended, rows_optimized,
                  maut_known, country
    """
    country = country.upper()
    zip_rate_map = pp.country_pallet_data(parsed, country)
    if not zip_rate_map:
        raise ValueError(f"No pallet rate data for {country} in the factor file.")
    bands = pp.band_ceilings(parsed)
 
    rows, maut_known = build_rows_pallet(country, zip_rate_map, bands,
                                         globals_, maut_table)
    df = pd.DataFrame(rows)
 
    out = Path(output_dir)
    ext_path = out / f'{country}_Pallet_extended.xlsx'
    opt_path = out / f'{country}_Pallet_optimized.xlsx'
 
    write_pallet_excel(df, ext_path, country)
    df_opt = optimize_pallet(df) if optimize else df
    write_pallet_excel(df_opt, opt_path, country)
 
    if not maut_known:
        log.warning("MAUT %% unknown for %s — MAUT column is 0. Fill MAUT_PCT['%s'].",
                    country, country)
 
    return {
        'country': country,
        'extended': str(ext_path),
        'optimized': str(opt_path),
        'rows_extended': len(df),
        'rows_optimized': len(df_opt),
        'maut_known': maut_known,
    }
