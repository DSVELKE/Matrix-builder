"""
master_parser.py — parse the DSV "MDK - FENDER - PARCEL RATES" master workbook.

This is the single source-of-truth rate card containing ALL carriers and ALL
countries in one file (one sheet per rate table). It is structured very
differently from the per-country files, so it gets its own parser.

Public API
----------
is_master_file(path)            -> bool      detect the master format
parse_master_rate_card(path)    -> dict      parse everything once
country_rate_data(master, iso2) -> dict      builder-compatible parsed dict
country_maut(master, iso2)      -> dict      {'DHL-ROS': pct, 'DPD': pct}
available_countries(master)     -> set       ISO2 codes with any rate data
"""

import re
import logging

import openpyxl

import pipeline as pl   # reuse _norm, _parse_float, _extract_tiers, etc.

log = logging.getLogger(__name__)

_ISO2 = re.compile(r'[A-Z]{2}')
_ZONE_KEY = re.compile(r'[A-Z]{2,3}\d*')   # AT, FR, ES4, ES5, ES6 …


# ==============================================================================
# Low-level helpers (whole-sheet scanning — master tables sit far from titles)
# ==============================================================================

def _scan_from_to(ws):
    """Find the From/To header anywhere in the sheet."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if pl._norm(ws.cell(r, c).value) in pl._FROM_WORDS:
                for off in (1, 2):
                    if pl._norm(ws.cell(r, c + off).value) in pl._TO_WORDS:
                        return r, c, c + off
    return None, None, None


def _find_label_row(ws, *labels):
    """Return (row, {normalised_label: col}) for the first row containing ALL labels."""
    needles = [pl._norm(l) for l in labels]
    for r in range(1, ws.max_row + 1):
        found = {}
        for c in range(1, ws.max_column + 1):
            cv = pl._norm(ws.cell(r, c).value)
            for n in needles:
                if cv == n:
                    found[n] = c
        if len(found) == len(needles):
            return r, found
    return None, {}


def _extract_zone_tiers(ws, hrow, from_col, to_col):
    """Extract {zone_key: tiers} for every zone column right of 'To'.
    Permissive: master tables sit alone on a sheet, so accept alphanumeric
    zone keys (ES4/ES5/ES6) and skip gaps rather than stopping."""
    by_zone = {}
    for c in range(to_col + 1, ws.max_column + 1):
        v = ws.cell(hrow, c).value
        if v is None:
            continue
        if isinstance(v, (int, float)):
            tiers = pl._extract_tiers(ws, hrow, from_col, to_col, c)
            if tiers:
                by_zone[int(v)] = tiers
        elif isinstance(v, str) and _ZONE_KEY.fullmatch(v.strip()):
            tiers = pl._extract_tiers(ws, hrow, from_col, to_col, c)
            if tiers:
                by_zone[v.strip()] = tiers
    return by_zone


def _flat_country_rates(ws, value_col_offset=1):
    """For a sheet listing 'XX  rate' rows (e.g. WEA, 7R9W62), return {ISO2: rate}.
    Scans every column for a 2-letter code followed by a numeric in the next col."""
    out = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and _ISO2.fullmatch(v.strip()):
                try:
                    out[v.strip()] = pl._parse_float(ws.cell(r, c + value_col_offset).value)
                    break
                except (TypeError, ValueError):
                    continue
    return out


# ==============================================================================
# Detection
# ==============================================================================

def is_master_file(path):
    """True if this looks like the DSV master workbook (one sheet per table)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return False
    names = wb.sheetnames
    if 'Cover Page' in names or 'Contents' in names:
        return True
    return sum(1 for n in names if n.upper().startswith('PARCEL -')) >= 3


# ==============================================================================
# Per-sheet parsers
# ==============================================================================

def _parse_upsde_zones(ws):
    """ZONES UPSDE -> {ISO2: [{'pc_from':int, 'pc_to':int, 'STDS':z, ...}, ...]}.

    Returns a LIST per country so all postcode ranges are preserved.
    Italy has 3 zones (00000-09999, 10000-50999, 51000-99999); the old design
    used out[iso] = entry which overwrote on every row, keeping only the last.
    Handles merged country cells by carrying the last seen ISO2 forward.
    """
    hrow, from_col, to_col = _scan_from_to(ws)
    if not hrow:
        return {}
    country_col = from_col - 1
    svc_cols = {}
    for label_row in (hrow, hrow - 1):
        for c in range(to_col + 1, ws.max_column + 1):
            vl = pl._norm(ws.cell(label_row, c).value)
            if 'standard single' in vl and 'STDS' not in svc_cols:
                svc_cols['STDS'] = c
            elif 'standard multi' in vl and 'STDM' not in svc_cols:
                svc_cols['STDM'] = c
            elif 'express saver' in vl and 'EXPRESS_SAVER' not in svc_cols:
                svc_cols['EXPRESS_SAVER'] = c
        if svc_cols:
            break

    out = {}
    last_iso = None
    for r in range(hrow + 1, ws.max_row + 1):
        country_val = ws.cell(r, country_col).value
        if isinstance(country_val, str) and _ISO2.fullmatch(country_val.strip()):
            last_iso = country_val.strip().upper()
        if last_iso is None:
            continue
        # Accept either numeric postcode ranges OR the literal token 'ALL'
        # (meaning "the entire country"). The per-country parser in pipeline.py
        # already maps ALL -> 0..99999; the master parser must do the same or
        # every country whose zone row says 'ALL' (~210 countries) gets dropped.
        pc_from_raw = ws.cell(r, from_col).value
        pc_to_raw   = ws.cell(r, to_col).value
        pc_prefix = None
        if isinstance(pc_from_raw, str) and pc_from_raw.strip().upper() == 'ALL':
            pc_from, pc_to = 0, 99999
        else:
            try:
                pc_from = int(str(pc_from_raw))
                pc_to   = int(str(pc_to_raw))
            except (TypeError, ValueError):
                # Alphanumeric postcode AREA (UK outward codes: AB, BT, WC, …).
                # These used to be dropped via `continue`, silently losing the
                # ENTIRE GB tariff (same failure mode as the old 'ALL' bug).
                # Keep the area string as a prefix; numeric pc range is N/A.
                if isinstance(pc_from_raw, str) and pc_from_raw.strip():
                    pc_prefix = pc_from_raw.strip().upper()
                    pc_from = pc_to = None
                else:
                    continue
        entry = {'pc_from': pc_from, 'pc_to': pc_to}
        if pc_prefix is not None:
            entry['pc_prefix'] = pc_prefix
        n_zone = 0
        for svc, col in svc_cols.items():
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                entry[svc] = int(v); n_zone += 1
            elif isinstance(v, str) and v.strip() and pl._norm(v) != 'on request':
                entry[svc] = v.strip(); n_zone += 1
        if n_zone:  # keep only rows with at least one real zone assignment
            out.setdefault(last_iso, []).append(entry)
    return out


def _parse_upsnl_zones(ws):
    """ZONES UPSNL EXPRESS -> {ISO2: zone_int} (Express Saver column)."""
    hrow, from_col, to_col = _scan_from_to(ws)
    if not hrow:
        return {}
    country_col = from_col - 1
    zone_col = None
    for label_row in (hrow, hrow - 1):
        for c in range(to_col + 1, ws.max_column + 1):
            if pl._norm(ws.cell(label_row, c).value) == 'express saver':
                zone_col = c
                break
        if zone_col:
            break
    out = {}
    if zone_col:
        for r in range(hrow + 1, ws.max_row + 1):
            country = ws.cell(r, country_col).value
            if not (isinstance(country, str) and _ISO2.fullmatch(country.strip())):
                continue
            try:
                out[country.strip().upper()] = int(str(ws.cell(r, zone_col).value))
            except (TypeError, ValueError):
                continue
    return out


def _parse_dpd(ws):
    """PARCEL - DPD -> {ISO2: {'normal': rate, 'small': rate}}."""
    hrow, cols = _find_label_row(ws, 'To Country')
    out = {}
    if not hrow:
        return out
    cc = cols[pl._norm('To Country')]
    for r in range(hrow + 1, ws.max_row + 1):
        country = ws.cell(r, cc).value
        if not (isinstance(country, str) and _ISO2.fullmatch(country.strip())):
            continue
        def f(col):
            try:
                return pl._parse_float(ws.cell(r, col).value)
            except (TypeError, ValueError):
                return None
        out[country.strip().upper()] = {'normal': f(cc + 1), 'small': f(cc + 2)}
    return out


def _parse_maut(ws):
    """MAUT SURCHARGE -> {ISO2: {'DHL-ROS': pct, 'DPD': pct}}.
    'On Request' / blank -> 0.0."""
    hrow, cols = _find_label_row(ws, 'To Country')
    out = {}
    if not hrow:
        return out
    cc = cols[pl._norm('To Country')]
    for r in range(hrow + 1, ws.max_row + 1):
        country = ws.cell(r, cc).value
        if not (isinstance(country, str) and _ISO2.fullmatch(country.strip())):
            continue
        def pct(col):
            try:
                return pl._parse_float(ws.cell(r, col).value)
            except (TypeError, ValueError):
                return 0.0
        out[country.strip().upper()] = {'DHL-ROS': pct(cc + 1), 'DPD': pct(cc + 2)}
    return out


def _parse_dhl_other(ws):
    """PARCEL - DHL - Other countries -> {ISO2: tiers}."""
    hrow, from_col, to_col = _scan_from_to(ws)
    out = {}
    if not hrow:
        return out
    for c in range(to_col + 1, ws.max_column + 1):
        v = ws.cell(hrow, c).value
        if isinstance(v, str) and _ISO2.fullmatch(v.strip()):
            tiers = pl._extract_tiers(ws, hrow, from_col, to_col, c)
            if tiers:
                out[v.strip().upper()] = tiers
    return out


def _parse_dhl_bnl(ws):
    """PARCEL - DHL - BNL -> {ISO2: {'first': rate, 'after': rate}}."""
    out = {}
    country_row = None
    country_cols = {}
    for r in range(1, ws.max_row + 1):
        codes = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)
                 if isinstance(ws.cell(r, c).value, str)
                 and _ISO2.fullmatch(str(ws.cell(r, c).value).strip())}
        if len(codes) >= 2:
            country_row, country_cols = r, codes
            break
    if country_row:
        first_row, after_row = country_row + 1, country_row + 2
        for col, code in country_cols.items():
            try:
                out[code.strip().upper()] = {
                    'first': pl._parse_float(ws.cell(first_row, col).value),
                    'after': pl._parse_float(ws.cell(after_row, col).value),
                }
            except (TypeError, ValueError):
                continue
    return out


def _parse_postnord(ws):
    """PARCEL - POSTNORD - STD -> {ISO2: {'B2B': r, 'HOME': r, 'PUDO': r}}."""
    out = {}
    country_row = None
    country_cols = {}
    for r in range(1, ws.max_row + 1):
        codes = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)
                 if isinstance(ws.cell(r, c).value, str)
                 and _ISO2.fullmatch(str(ws.cell(r, c).value).strip())}
        if len(codes) >= 2:
            country_row, country_cols = r, codes
            break
    if country_row:
        svc_col = min(country_cols) - 1
        for r in range(country_row + 1, ws.max_row + 1):
            svc_raw = ws.cell(r, svc_col).value
            if not isinstance(svc_raw, str):
                continue
            svc = svc_raw.split('|')[0].strip().upper()
            if not svc:
                continue
            for col, code in country_cols.items():
                try:
                    rate = pl._parse_float(ws.cell(r, col).value)
                    out.setdefault(code.strip().upper(), {})[svc] = rate
                except (TypeError, ValueError):
                    continue
    return out


def _parse_gb_tiers(ws):
    """Single-rate-column tier table (UPSGB STDS/STDM/EXPS)."""
    hrow, from_col, to_col = _scan_from_to(ws)
    if not hrow:
        return []
    return pl._extract_tiers(ws, hrow, from_col, to_col, to_col + 1)


def _parse_linehaul(ws):
    """First numeric value next to a Carrier/UPS label."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and ('UPS' in v.upper()):
                nxt = ws.cell(r, c + 1).value
                if isinstance(nxt, (int, float)):
                    return float(nxt)
    return None


# ==============================================================================
# Master parse
# ==============================================================================

def parse_master_rate_card(path):
    """Parse the entire master workbook once into a structured dict."""
    wb = openpyxl.load_workbook(path, data_only=True)

    def sheet(*hints):
        return pl._find_sheet(wb, *hints)

    master = {
        'UPSDE': {}, 'UPSNL': {}, 'DHL': {}, 'DPD': {},
        'POSTNORD': {}, 'MAUT': {}, 'UPSGB': {},
    }

    # ── UPSDE ────────────────────────────────────────────────────────────────
    ws = sheet('ZONES UPSDE')
    if ws:
        master['UPSDE']['zones_by_country'] = _parse_upsde_zones(ws)
    for key, hint in [('STDS_by_zone', 'PARCEL - UPS - STDS'),
                      ('STDM_by_zone', 'PARCEL - UPS - STDM'),
                      ('EXPRESS_SAVER_by_zone', 'PARCEL - EXPRESS SAVER UPSDE')]:
        ws = sheet(hint)
        if ws:
            hrow, fc, tc = _scan_from_to(ws)
            if hrow:
                master['UPSDE'][key] = _extract_zone_tiers(ws, hrow, fc, tc)
    ws = sheet('PARCEL - EXPSAVER UPSDE 7R9W62')
    if ws:
        master['UPSDE']['expsaver_7r9w62'] = _flat_country_rates(ws)
    ws = sheet('PARCEL - UPS - WEA')
    if ws:
        master['UPSDE']['wea'] = _flat_country_rates(ws)
    ws = sheet('PARCEL - UPS DE - LINEHAUL', 'PARCEL - UPS - LINEHAUL')
    if ws:
        master['UPSDE']['linehaul'] = _parse_linehaul(ws)

    # ── UPSNL ────────────────────────────────────────────────────────────────
    ws = sheet('ZONES UPSNL EXPRESS', 'ZONES UPSNL')
    if ws:
        master['UPSNL']['zones_by_country'] = _parse_upsnl_zones(ws)
    ws = sheet('PARCEL - EXPRESS SAVER UPSNL')
    if ws:
        hrow, fc, tc = _scan_from_to(ws)
        if hrow:
            rbz = _extract_zone_tiers(ws, hrow, fc, tc)
            master['UPSNL']['rates_by_zone'] = {k: v for k, v in rbz.items()
                                                if isinstance(k, int)}

    # ── DHL ──────────────────────────────────────────────────────────────────
    ws = sheet('PARCEL - DHL - Other countries')
    if ws:
        master['DHL']['other'] = _parse_dhl_other(ws)
    ws = sheet('PARCEL - DHL - BNL')
    if ws:
        master['DHL']['bnl'] = _parse_dhl_bnl(ws)

    # ── DPD ──────────────────────────────────────────────────────────────────
    ws = sheet('PARCEL - DPD')
    if ws:
        master['DPD'] = _parse_dpd(ws)

    # ── PostNord ──────────────────────────────────────────────────────────────
    ws = sheet('PARCEL - POSTNORD - STD', 'PARCEL - POSTNORD')
    if ws:
        master['POSTNORD'] = _parse_postnord(ws)

    # ── MAUT ──────────────────────────────────────────────────────────────────
    ws = sheet('MAUT SURCHARGE', 'MAUT')
    if ws:
        master['MAUT'] = _parse_maut(ws)

    # ── UPSGB ──────────────────────────────────────────────────────────────────
    gb = {}
    ws = sheet('PARCEL - UPSGB - STDS')
    if ws: gb['STDS'] = _parse_gb_tiers(ws)
    ws = sheet('PARCEL - UPSGB - STDM')
    if ws: gb['STDM'] = _parse_gb_tiers(ws)
    ws = sheet('PARCEL - UPSGB - EXPS')
    if ws: gb['EXPS'] = _parse_gb_tiers(ws)
    ws = sheet('PARCEL - UPS GB - LINEHAUL', 'PARCEL - UPSGB - LINEHAUL')
    if ws: gb['linehaul'] = _parse_linehaul(ws)
    if gb:
        master['UPSGB'] = gb

    return master


# ==============================================================================
# Country extraction → builder-compatible dict
# ==============================================================================

def country_rate_data(master, iso2):
    """Return a parsed dict in the same shape parse_rate_cards() produces,
    containing only the data relevant to `iso2`."""
    iso2 = iso2.upper()
    parsed = {}

    # ── UPDE ──────────────────────────────────────────────────────────────────
    upsde = master.get('UPSDE', {})
    zone_rows = upsde.get('zones_by_country', {}).get(iso2, [])
    upde = {}
    if zone_rows:
        # Each row is {'pc_from':int, 'pc_to':int, 'STDS':z, 'STDM':z, ...}
        upde['zones'] = [{'country': iso2, **row} for row in zone_rows]
        upde['STDS_by_zone'] = upsde.get('STDS_by_zone', {})
        upde['STDM_by_zone'] = upsde.get('STDM_by_zone', {})
        upde['EXPRESS_SAVER_by_zone'] = upsde.get('EXPRESS_SAVER_by_zone', {})
    if iso2 in upsde.get('expsaver_7r9w62', {}):
        upde['EXPSAVER_7R9W62'] = upsde['expsaver_7r9w62'][iso2]
    if iso2 in upsde.get('wea', {}):
        upde['WEA'] = upsde['wea'][iso2]
    if upde:
        parsed['UPDE'] = upde

    # ── UPSNL ──────────────────────────────────────────────────────────────────
    upsnl = master.get('UPSNL', {})
    zone = upsnl.get('zones_by_country', {}).get(iso2)
    if zone is not None and upsnl.get('rates_by_zone'):
        parsed['UPSNL'] = {
            'zones': [{'country': iso2, 'pc_from': 0, 'pc_to': 99999, 'zone': zone}],
            'rates_by_zone': upsnl['rates_by_zone'],
        }

    # ── DHL ────────────────────────────────────────────────────────────────────
    dhl = master.get('DHL', {})
    if iso2 in dhl.get('bnl', {}):
        parsed['DHL-ROS'] = {'bnl': dhl['bnl'][iso2]}
    elif iso2 in dhl.get('other', {}):
        parsed['DHL-ROS'] = {'STANDARD': dhl['other'][iso2]}

    # ── DPD ────────────────────────────────────────────────────────────────────
    if iso2 in master.get('DPD', {}):
        d = master['DPD'][iso2]
        dpd = {}
        if d.get('normal') is not None:
            dpd['groot'] = d['normal']
        if d.get('small') is not None:
            dpd['klein'] = d['small']
        if dpd:
            parsed['DPD'] = dpd

    # ── PostNord ────────────────────────────────────────────────────────────────
    if iso2 in master.get('POSTNORD', {}):
        parsed['POSTNORD'] = {'flat_rates': master['POSTNORD'][iso2]}

    # ── UPSGB (GB only) ──────────────────────────────────────────────────────────
    if iso2 == 'GB' and master.get('UPSGB'):
        parsed['UPSGB'] = dict(master['UPSGB'])

    return parsed


def country_maut(master, iso2):
    """Return {'DHL-ROS': pct, 'DPD': pct} for the country (0.0 if absent)."""
    m = master.get('MAUT', {}).get(iso2.upper(), {})
    return {'DHL-ROS': m.get('DHL-ROS', 0.0), 'DPD': m.get('DPD', 0.0)}


def available_countries(master):
    """All ISO2 codes that have ANY rate data in the master file."""
    countries = set()
    countries |= set(master.get('UPSDE', {}).get('zones_by_country', {}))
    countries |= set(master.get('UPSNL', {}).get('zones_by_country', {}))
    countries |= set(master.get('DHL', {}).get('other', {}))
    countries |= set(master.get('DHL', {}).get('bnl', {}))
    countries |= set(master.get('DPD', {}))
    countries |= set(master.get('POSTNORD', {}))
    if master.get('UPSGB'):
        countries.add('GB')
    return countries
