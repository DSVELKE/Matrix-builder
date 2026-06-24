"""
pipeline.py — Rate Matrix Builder core module.

All logic lives here. The Streamlit app (app.py) imports and calls run_pipeline().
Key design decision: every function that reads CARRIER_DEFAULTS or VARIABLES_LAYOUT
accepts an optional override so the app can customise rates per session without
touching module-level globals (important for concurrent Streamlit users).
"""

import re
import logging
import shutil
from copy import deepcopy
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

log = logging.getLogger(__name__)


# ==============================================================================
# 1. CONFIGURATION
# ==============================================================================

CARRIER_DEFAULTS = {
    'UPDE': {
        'label': 'UPS DE',
        'services': ['STANDARD', 'EXPRESS SAVER 7R9W62', 'EXPRESS SAVER'],
        'has_postcode': True,
        'volume_divisor': 167,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B1',
        'maut_pct': 0.0,
        'maut_variables_ref': None,
        'linehaul_per_parcel': 1.31,
    },
    'UPSNL': {
        'label': 'UPS NL',
        'services': ['EXPRESS SAVER'],
        'has_postcode': True,
        'volume_divisor': 167,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B4',
        'maut_pct': 0.0,
        'maut_variables_ref': None,
        'linehaul_per_parcel': 0.0,
    },
    'DPD': {
        'label': 'DPD',
        'services': ['PARCEL'],
        'has_postcode': False,
        'volume_divisor': 250,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B3',
        'maut_pct': 0.05,
        'maut_variables_ref': 'B8',
        'linehaul_per_parcel': 0.0,
    },
    'DHL-ROS': {
        'label': 'DHL',
        'services': ['STANDARD'],
        'has_postcode': False,
        'volume_divisor': 250,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B2',
        'maut_pct': 0.06,
        'maut_variables_ref': 'B9',
        'linehaul_per_parcel': 0.0,
    },
    'POSTNORD': {
        'label': 'PostNord',
        'services': ['STANDARD'],
        'has_postcode': False,
        'volume_divisor': 250,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B5',
        'maut_pct': 0.0,
        'maut_variables_ref': None,
        'linehaul_per_parcel': 0.0,
    },
    'UPSGB': {
        'label': 'UPS GB',
        'services': ['STANDARD', 'EXPRESS SAVER'],
        'has_postcode': False,
        'volume_divisor': 167,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B6',
        'maut_pct': 0.0,
        'maut_variables_ref': None,
        'linehaul_per_parcel': 3.9,
    },
}

VARIABLES_LAYOUT = [
    ('FUEL UPSDE',    0.27),   # B1
    ('FUEL DHL',      0.27),   # B2
    ('FUEL DPD',      0.27),   # B3
    ('FUEL UPSNL',    0.27),   # B4
    ('FUEL POSTNORD', 0.27),   # B5
    ('FUEL UPSGB',    0.27),   # B6
    (None, None),              # B7  blank spacer
    ('MAUT DPD',      0.05),   # B8
    ('MAUT DHL',      0.06),   # B9
]

_BASE_CARRIERS   = ['UPDE', 'DPD', 'DHL-ROS', 'UPSNL']
_SCANDI_CARRIERS = ['UPDE', 'DPD', 'DHL-ROS', 'UPSNL', 'POSTNORD']
_SCANDI_ISO      = {'SE', 'DK', 'NO', 'FI'}


def _default_country_cfg(iso2):
    return {
        'iso2':                  iso2,
        'site_id':               'NLMOE01',
        'client_id':             'NLFENDER',
        'max_parcel_count':      15,
        'max_each_weight_kg':    31.5,
        'each_weight_grid':      sorted(set(list(range(1, 32)) + [31.5])),
        'carriers':              _SCANDI_CARRIERS if iso2 in _SCANDI_ISO else _BASE_CARRIERS,
        'postcode_prefix_range': (0, 99),
    }


COUNTRY_CONFIG = {iso: _default_country_cfg(iso) for iso in [
    'DE', 'FR', 'IT', 'ES', 'NL',
    'BE', 'IE', 'PT', 'LU',
    'AT', 'CH', 'PL', 'CZ', 'SK', 'HU', 'SI',
    'SE', 'DK', 'NO', 'FI',
    'GR', 'HR', 'BG', 'RO', 'SM',
    'EE', 'LV', 'LT', 'LI',
]}
# DE only has three carriers (no UPSNL)
COUNTRY_CONFIG['DE']['carriers'] = ['UPDE', 'DPD', 'DHL-ROS']

# GB — UK domestic via UPSGB, plus NL-origin export carriers that quote GB
COUNTRY_CONFIG['GB'] = _default_country_cfg('GB')
COUNTRY_CONFIG['GB']['carriers'] = ['UPSGB', 'UPDE', 'DPD', 'DHL-ROS', 'UPSNL']


# ==============================================================================
# 2. ROBUST TEXT / SHEET HELPERS
# ==============================================================================

_JUNK = re.compile(r'[\s\-_/\\.,;:()\[\]]+')


def _norm(v):
    """Normalise any value: lowercase, collapse all punctuation/whitespace to ' '."""
    if v is None:
        return ''
    return _JUNK.sub(' ', str(v).lower()).strip()


def _cell_match(cell_value, *needles):
    """True if the normalised cell value contains/equals ANY normalised needle."""
    cv = _norm(cell_value)
    if not cv:
        return False
    for raw in needles:
        n = _norm(raw)
        if n and (n in cv or cv in n):
            return True
    return False


def _parse_float(v):
    """Parse float robustly: handles comma-decimals, currency symbols, None."""
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        raise ValueError('None')
    s = re.sub(r'[€$£\s]', '', str(v).strip())
    if ',' in s and '.' in s:
        if s.index(',') > s.index('.'):        # '1.234,56' European
            s = s.replace('.', '').replace(',', '.')
        else:                                   # '1,234.56' Anglo
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
    return float(s)


def _find_sheet(wb, *name_hints):
    """Return first sheet whose name fuzzy-matches any hint; None if not found."""
    for hint in name_hints:
        if hint in wb.sheetnames:
            return wb[hint]
    for hint in name_hints:
        hn = _norm(hint)
        for sname in wb.sheetnames:
            if hn and (_norm(sname) == hn or hn in _norm(sname)):
                return wb[sname]
    return None


def _scan_anchor(ws, *anchor_texts, max_row=None, max_col=None):
    """First cell that fuzzy-matches any anchor text; None if not found."""
    mr = max_row or ws.max_row
    mc = max_col or ws.max_column
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            v = ws.cell(r, c).value
            if v is not None and _cell_match(v, *anchor_texts):
                return (r, c)
    return None


# keep original name as alias for backward compat
_scan_for_anchor = _scan_anchor

_FROM_WORDS = {'from', 'van', 'von', 'de', 'fra', 'vanaf', 'weight from', 'kg from'}
_TO_WORDS   = {'to', 'tot', 'bis', 'a', 'til', 'tot en met', 'weight to', 'kg to'}


def _find_from_to(ws, anchor_row, anchor_col, max_rows=10, col_slack=5):
    """Find From/To header row near anchor. Multi-language, wider search window."""
    col_lo = max(1, anchor_col - col_slack)
    col_hi = min(anchor_col + col_slack, ws.max_column)
    row_lo = max(1, anchor_row - 1)
    row_hi = min(anchor_row + max_rows, ws.max_row)

    for r in range(row_lo, row_hi + 1):
        for c in range(col_lo, col_hi + 1):
            cv = _norm(ws.cell(r, c).value)
            if cv not in _FROM_WORDS:
                continue
            for to_offset in (1, 2):
                if _norm(ws.cell(r, c + to_offset).value) in _TO_WORDS:
                    return r, c, c + to_offset
    return None, None, None


_find_from_to_header = _find_from_to   # alias

_OVER_RE = re.compile(
    r'\b(over|meer|plus|above|mehr|oltre|mas|vidare|sup[ée]rieur)\b|\+\s*$'
)


def _extract_tiers(ws, header_row, from_col, to_col, rate_col, max_rows=200):
    """Extract weight-band tiers; tolerates blank spacers, comma-decimals,
    all 'over/meer/plus' end markers."""
    tiers = []
    last_from = -1
    blanks = 0

    for r in range(header_row + 1, min(ws.max_row, header_row + max_rows) + 1):
        f_raw = ws.cell(r, from_col).value
        t_raw = ws.cell(r, to_col).value
        rate_raw = ws.cell(r, rate_col).value

        if f_raw is None and rate_raw is None:
            blanks += 1
            if blanks > 2:
                break
            continue
        blanks = 0

        try:
            f_val = _parse_float(f_raw)
        except (TypeError, ValueError):
            break

        if tiers and f_val < last_from and f_val <= 1:
            break
        last_from = f_val

        try:
            rate_val = _parse_float(rate_raw)
        except (TypeError, ValueError):
            continue

        t_str = _norm(t_raw) if t_raw is not None else ''
        if t_raw is None or _OVER_RE.search(t_str):
            tiers.append({'from': f_val, 'to': float('inf'),
                          'rate': rate_val, 'per_kg': True})
            break
        try:
            t_val = _parse_float(t_raw)
        except (TypeError, ValueError):
            break

        tiers.append({'from': f_val, 'to': t_val, 'rate': rate_val, 'per_kg': False})

    return tiers


_extract_tier_table = _extract_tiers   # alias


def _extract_rates_by_zone(ws, hrow, from_col, to_col):
    by_zone = {}
    for c in range(to_col + 1, ws.max_column + 1):
        v = ws.cell(hrow, c).value
        if v is None:
            if by_zone:
                break
            continue
        if isinstance(v, (int, float)):
            tiers = _extract_tiers(ws, hrow, from_col, to_col, c)
            if tiers:
                by_zone[int(v)] = tiers
            continue
        if isinstance(v, str):
            vs = v.strip()
            if _norm(vs) in ('from', 'to', 'payweight', 'van', 'tot'):
                break
            if re.fullmatch(r'[A-Z]{2,3}', vs):
                tiers = _extract_tiers(ws, hrow, from_col, to_col, c)
                if tiers:
                    by_zone[vs] = tiers
                continue
            if by_zone:
                break
    return by_zone


# ==============================================================================
# 3. RATE-CARD PARSER
# ==============================================================================

def _parse_upsde_zones(ws):
    anchor = _scan_anchor(ws, 'Zones UPSDE', 'ZONES UPSDE', 'zones ups de')
    if not anchor:
        return []
    ar, ac = anchor
    hrow, from_col, to_col = _find_from_to(ws, ar, ac, max_rows=4, col_slack=4)
    if not hrow:
        return []
    country_col = from_col - 1
    service_cols = {}
    for label_row in (hrow, hrow - 1):
        if label_row < 1:
            continue
        for c in range(to_col + 1, ws.max_column + 1):
            v = ws.cell(label_row, c).value
            if not isinstance(v, str):
                continue
            vl = _norm(v)
            if 'standard single' in vl and 'STDS' not in service_cols:
                service_cols['STDS'] = c
            elif 'standard multi' in vl and 'STDM' not in service_cols:
                service_cols['STDM'] = c
            elif 'express saver' in vl and 'EXPRESS_SAVER' not in service_cols:
                service_cols['EXPRESS_SAVER'] = c
        if service_cols:
            break

    zones = []
    for r in range(hrow + 1, ws.max_row + 1):
        country = ws.cell(r, country_col).value
        pc_from = ws.cell(r, from_col).value
        pc_to   = ws.cell(r, to_col).value
        if country is None and pc_from is None and pc_to is None:
            break
        if isinstance(pc_from, str) and pc_from.strip().upper() == 'ALL':
            pc_from_int, pc_to_int = 0, 99999
        else:
            try:
                pc_from_int = int(str(pc_from))
                pc_to_int   = int(str(pc_to)) if pc_to is not None else pc_from_int
            except (TypeError, ValueError):
                continue
        entry = {'country': country, 'pc_from': pc_from_int, 'pc_to': pc_to_int}
        for svc, col in service_cols.items():
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                entry[svc] = int(v)
            elif isinstance(v, str) and v.strip():
                entry[svc] = v.strip()
        zones.append(entry)
    return zones


# PostNord service anchors
_PN_SHEET_NAMES   = ['POSTNORD', 'POST NORD', 'PostNord', 'Post Nord', 'PN']
_PN_SVC_ANCHORS   = {
    'STANDARD': ['parcel postnord', 'postnord standard', 'postnord parcel',
                 'rate per parcel postnord', 'standard postnord', 'postnord'],
    'EXPRESS':  ['postnord express', 'express postnord'],
    'ECONOMY':  ['postnord economy', 'economy postnord'],
}


def _parse_postnord_sheet(ws):
    """
    Four-strategy PostNord parser.

    Strategy 0 — Flat rate per service code (e.g. SE: B2B|18|15P → 12.2)
        Looks for a 'Servicelevel' header; collects service-code rows where
        the right column is numeric.  Returns 'flat_rates': {name: rate}.

    Strategy 1 — Anchored service tier tables (From/To weight bands)
    Strategy 2 — Full-sheet From/To scan
    Strategy 3 — Numeric region detection
    """
    result = {}

    # ── Strategy 0: flat rate per service code ────────────────────────────────
    anchor = _scan_anchor(ws, 'servicelevel', 'service level', 'service code')
    if anchor:
        ar, ac = anchor
        for svc_col in (ac, ac + 1):
            flat_rates = {}
            for r in range(ar + 1, ws.max_row + 1):
                svc_raw = ws.cell(r, svc_col).value
                if svc_raw is None:
                    continue
                svc_str = str(svc_raw).strip()
                # Skip pure country-code rows like 'SE', 'DK'
                if re.fullmatch(r'[A-Z]{2,3}', svc_str):
                    continue
                # Skip header-like strings
                if _norm(svc_str) in ('to country', 'country', 'to', 'rate',
                                      'service', 'servicelevel'):
                    continue
                for rate_offset in (1, 2):
                    try:
                        rate = _parse_float(ws.cell(r, svc_col + rate_offset).value)
                        svc_name = svc_str.split('|')[0].strip().upper()
                        if svc_name:
                            flat_rates[svc_name] = rate
                        break
                    except (TypeError, ValueError):
                        continue
            if flat_rates:
                result['flat_rates'] = flat_rates
                log.info('    PostNord strategy 0: flat rates %s', flat_rates)
                return result

    # ── Strategy 1: anchored service tier tables ──────────────────────────────
    for svc, anchors in _PN_SVC_ANCHORS.items():
        anchor = _scan_anchor(ws, *anchors)
        if not anchor:
            continue
        ar, ac = anchor
        hrow, fc, tc = _find_from_to(ws, ar, ac, max_rows=12, col_slack=6)
        if not hrow:
            continue
        by_zone = _extract_rates_by_zone(ws, hrow, fc, tc)
        if by_zone:
            result[f'{svc}_by_zone'] = by_zone
            result[svc] = next(iter(by_zone.values()))
        else:
            tiers = _extract_tiers(ws, hrow, fc, tc, tc + 1)
            if tiers:
                result[svc] = tiers
        if svc in result:
            break

    if result:
        return result

    # ── Strategy 2: full-sheet From/To scan ──────────────────────────────────
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if _norm(ws.cell(r, c).value) not in _FROM_WORDS:
                continue
            for off in (1, 2):
                if _norm(ws.cell(r, c + off).value) not in _TO_WORDS:
                    continue
                tiers = _extract_tiers(ws, r, c, c + off, c + off + 1)
                if len(tiers) >= 2:
                    result['STANDARD'] = tiers
                    return result

    # ── Strategy 3: numeric region detection ─────────────────────────────────
    for sc in range(1, max(1, ws.max_column - 2) + 1):
        for sr in range(1, ws.max_row + 1):
            run = 0
            for rr in range(sr, min(sr + 30, ws.max_row) + 1):
                try:
                    _parse_float(ws.cell(rr, sc).value)
                    _parse_float(ws.cell(rr, sc + 1).value)
                    _parse_float(ws.cell(rr, sc + 2).value)
                    run += 1
                except (TypeError, ValueError):
                    break
            if run >= 3:
                tiers = _extract_tiers(ws, sr - 1, sc, sc + 1, sc + 2)
                if len(tiers) >= 2:
                    result['STANDARD'] = tiers
                    return result

    log.warning('PostNord: all strategies failed on sheet "%s"', ws.title)
    return result


def parse_rate_cards(excel_path):
    """Parse all carrier rate tables from the uploaded Excel. Fuzzy sheet matching."""
    wb  = openpyxl.load_workbook(excel_path, data_only=True)
    out = {}

    # ── UPSDE ────────────────────────────────────────────────────────────────
    ws = _find_sheet(wb, 'UPSDE', 'UPS DE', 'UPS-DE', 'UPDE')
    if ws:
        upde = {'zones': _parse_upsde_zones(ws)}
        for label, key in [('PARCEL - UPS - STDS', 'STDS'),
                            ('PARCEL - UPS - STDM', 'STDM')]:
            anchor = _scan_anchor(ws, label)
            if anchor:
                ar, ac = anchor
                hrow, fc, tc = _find_from_to(ws, ar, ac)
                if hrow:
                    upde[f'{key}_by_zone'] = _extract_rates_by_zone(ws, hrow, fc, tc)
                    upde[key] = _extract_tiers(ws, hrow, fc, tc, tc + 1)

        anchor = _scan_anchor(ws, 'EXPSAVER UPSDE 7R9W62', 'EXPSAVER 7R9W62')
        if anchor:
            ar, ac = anchor
            for r in range(ar + 1, min(ar + 6, ws.max_row) + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and re.fullmatch(r'[A-Z]{2}', v.strip()):
                        try:
                            upde['EXPSAVER_7R9W62'] = _parse_float(ws.cell(r, c + 1).value)
                            break
                        except (TypeError, ValueError):
                            pass
                if 'EXPSAVER_7R9W62' in upde:
                    break

        anchor = _scan_anchor(ws, 'UPS DE - LINEHAUL', 'UPSDE Linehaul', 'linehaul ups')
        if anchor:
            ar, ac = anchor
            for r in range(ar + 1, min(ar + 6, ws.max_row) + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and any(
                        kw in v.upper() for kw in ('DEUTSCHLAND', 'GERMANY', 'DE')
                    ):
                        try:
                            upde['LINEHAUL'] = _parse_float(ws.cell(r, c + 1).value)
                            break
                        except (TypeError, ValueError):
                            pass
                if 'LINEHAUL' in upde:
                    break

        anchor = _scan_anchor(ws, 'EXPRESS SAVER UPSDE', 'PARCEL - EXPRESS SAVER UPSDE')
        if anchor:
            ar, ac = anchor
            hrow, fc, tc = _find_from_to(ws, ar, ac)
            if hrow:
                upde['EXPRESS_SAVER_by_zone'] = _extract_rates_by_zone(ws, hrow, fc, tc)
                upde['EXPRESS_SAVER'] = _extract_tiers(ws, hrow, fc, tc, tc + 1)

        out['UPDE'] = upde

    # ── UPSNL ────────────────────────────────────────────────────────────────
    ws = _find_sheet(wb, 'UPSNL', 'UPS NL', 'UPS-NL')
    if ws:
        upsnl = {'zones': [], 'rates_by_zone': {}}
        anchor = _scan_anchor(ws, 'ZONES UPSNL', 'Zones UPSNL')
        if anchor:
            ar, ac = anchor
            hrow, fc, tc = _find_from_to(ws, ar, ac, max_rows=5, col_slack=4)
            if hrow:
                country_col = fc - 1
                zone_col = None
                for c in range(tc + 1, ws.max_column + 1):
                    if _cell_match(ws.cell(hrow, c).value, 'express saver'):
                        zone_col = c
                        break
                if zone_col is None:
                    for c in range(tc + 1, ws.max_column + 1):
                        if isinstance(ws.cell(hrow + 1, c).value, (int, float)):
                            zone_col = c
                            break
                if zone_col:
                    for r in range(hrow + 1, ws.max_row + 1):
                        pc_from  = ws.cell(r, fc).value
                        pc_to    = ws.cell(r, tc).value
                        zone_raw = ws.cell(r, zone_col).value
                        if pc_from is None and pc_to is None and zone_raw is None:
                            break
                        if isinstance(pc_from, str) and pc_from.strip().upper() == 'ALL':
                            pc_from_int, pc_to_int = 0, 99999
                        else:
                            try:
                                pc_from_int = int(str(pc_from))
                                pc_to_int   = int(str(pc_to))
                            except (TypeError, ValueError):
                                continue
                        try:
                            upsnl['zones'].append({
                                'country':  ws.cell(r, country_col).value,
                                'pc_from':  pc_from_int,
                                'pc_to':    pc_to_int,
                                'zone':     int(str(zone_raw)),
                            })
                        except (TypeError, ValueError):
                            continue

        anchor = _scan_anchor(ws, 'PARCEL - EXPRESS SAVER UPSNL', 'EXPRESS SAVER UPSNL',
                              'Rates UPSNL express saver', 'Rates UPSNL')
        if anchor:
            ar, ac = anchor
            hrow, fc, tc = _find_from_to(ws, ar, ac, max_rows=4, col_slack=3)
            if hrow:
                rbz = _extract_rates_by_zone(ws, hrow, fc, tc)
                upsnl['rates_by_zone'] = {k: v for k, v in rbz.items()
                                          if isinstance(k, int)}
        out['UPSNL'] = upsnl

    # ── DHL ──────────────────────────────────────────────────────────────────
    ws = _find_sheet(wb, 'DHL', 'DHL-ROS', 'DHL ROS')
    if ws:
        anchor = _scan_anchor(ws, 'PARCEL - DHL', 'Rate per parcel DHL', 'dhl standard')
        if anchor:
            ar, ac = anchor
            hrow, fc, tc = _find_from_to(ws, ar, ac)
            if hrow:
                tiers = _extract_tiers(ws, hrow, fc, tc, tc + 1)
                if tiers:
                    out['DHL-ROS'] = {'STANDARD': tiers}

    # ── DPD ──────────────────────────────────────────────────────────────────
    ws = _find_sheet(wb, 'DPD')
    if ws:
        anchor = _scan_anchor(ws, 'PARCEL - DPD', 'Rate per parcel DPD', 'dpd parcel')
        if anchor:
            ar, _ = anchor
            label_map = {'groot': 'groot', 'klein': 'klein', 'big': 'groot',
                         'small': 'klein', 'large': 'groot', 'heavy': 'groot'}
            rates = {}
            for r in range(ar + 1, min(ar + 6, ws.max_row) + 1):
                row_labels = {}
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str):
                        k = label_map.get(v.strip().lower())
                        if k:
                            row_labels[c] = k
                if row_labels:
                    for col, norm_label in row_labels.items():
                        try:
                            rates[norm_label] = _parse_float(ws.cell(r - 1, col).value)
                        except (TypeError, ValueError):
                            pass
                    break
            if rates:
                out['DPD'] = rates

    # ── POSTNORD ─────────────────────────────────────────────────────────────
    ws = _find_sheet(wb, *_PN_SHEET_NAMES)
    if ws:
        data = _parse_postnord_sheet(ws)
        if data:
            out['POSTNORD'] = data
        else:
            log.warning('POSTNORD sheet found but no rates parsed')

    return out


# ==============================================================================
# 4. TIER UTILITIES
# ==============================================================================

def lookup_tier_rate(tiers, weight):
    if weight <= 0:
        return None
    for t in tiers:
        if t['from'] < weight <= t['to']:
            return t['rate'] * weight if t['per_kg'] else t['rate']
    if tiers and weight == tiers[0]['from']:
        return tiers[0]['rate']
    return None


def collapse_same_rate_tiers(tiers, weight_cap=None):
    if not tiers:
        return []
    bands, cur_rate, cur_to, cur_pk = [], tiers[0]['rate'], tiers[0]['to'], tiers[0].get('per_kg', False)
    for t in tiers[1:]:
        if t['rate'] == cur_rate and not cur_pk:
            cur_to = t['to']
        else:
            bands.append((cur_to, cur_rate, cur_pk))
            cur_rate, cur_to, cur_pk = t['rate'], t['to'], t.get('per_kg', False)
    bands.append((cur_to, cur_rate, cur_pk))
    if weight_cap is not None:
        bands = [(min(to, weight_cap), rate, pk) for to, rate, pk in bands
                 if to == float('inf') or to <= weight_cap * 1.5]
        bands = [(weight_cap if to == float('inf') else to, rate, pk)
                 for to, rate, pk in bands]
        seen, out = set(), []
        for to, rate, pk in bands:
            if (to, rate) not in seen:
                seen.add((to, rate))
                out.append((to, rate, pk))
        bands = out
    return bands


# ==============================================================================
# 5. MATRIX BUILDERS
# ==============================================================================

def _upde_service_buckets(rate_data, service_key, country_cfg):
    """Return [(postcode_prefix_or_None, tiers), ...] for a UPDE service.

    Postcode-resistance rules:
    - No zones at all            → flat rate, no postcode column
    - Zones but no pc_from/pc_to → treat as single zone, no postcode column
    - All zones map to same rate → collapse to single entry, no postcode column
    - Multiple distinct zones    → one entry per postcode prefix in range
    - Prefix not covered by any  → falls back to the worst (highest-rate) zone
      so no order is ever left unmatched
    """
    zones   = rate_data.get('zones', [])
    by_zone = rate_data.get(f'{service_key}_by_zone', {})
    flat    = rate_data.get(service_key, [])
    pc_min, pc_max = country_cfg['postcode_prefix_range']

    if not zones:
        return [(None, flat)] if flat else []

    zone_ids = [z[service_key] for z in zones if service_key in z]
    if not zone_ids:
        return [(None, flat)] if flat else []
    unique_zones = set(zone_ids)

    def tiers_for(zid):
        if isinstance(zid, int):
            return by_zone.get(zid, [])
        return by_zone.get(zid) or flat

    # Alphanumeric-postcode countries (UK): zones are keyed by outward-code AREA
    # (AB, BT, WC, …) instead of numeric prefixes. Emit one bucket per area so
    # each carries its true zone — Express Saver mainland z3 vs BT (N. Ireland)
    # z4 vs Channel Islands (GY/JE) z5. These deliberately do NOT collapse to a
    # blank country-wide bucket: a cheaper blank row would capture the dearer
    # BT/CI postcodes under CargoWrite cheapest-first matching and underprice
    # them. Areas whose value for this service is "On Request" (HS/ZE Standard)
    # simply have no zone key here and are skipped.
    if any(z.get('pc_prefix') for z in zones):
        return [(z['pc_prefix'], tiers_for(z[service_key]))
                for z in zones
                if service_key in z and tiers_for(z[service_key])]

    # If zones carry no pc_from/pc_to (old-style single-zone entry), treat as flat
    has_pc_ranges = any('pc_from' in z and 'pc_to' in z for z in zones)
    if not has_pc_ranges or len(unique_zones) == 1:
        return [(None, tiers_for(next(iter(unique_zones))))]

    # Find the fallback: worst zone (highest rate at max weight)
    def _zone_max_rate(zid):
        t = tiers_for(zid)
        return max((b['rate'] for b in t), default=0)
    fallback_zid = max(unique_zones, key=_zone_max_rate)

    buckets = []
    for pc in range(pc_min, pc_max + 1):
        pc_full = pc * 1000
        zid = next((z[service_key] for z in zones
                    if z.get('pc_from', 0) <= pc_full <= z.get('pc_to', 99999)
                    and service_key in z), fallback_zid)
        t = tiers_for(zid)
        if t:
            buckets.append((pc, t))

    # If all prefixes resolved to the same zone, collapse to no-postcode
    if buckets and len({t_id for _, t_id in
                        [(pc, id(t)) for pc, t in buckets]}) == 1:
        return [(None, buckets[0][1])]

    return buckets


def _common(site, client, carrier, iso2):
    return {'SITE_ID': site, 'CLIENT_ID': client, 'CARRIER_ID': carrier,
            'COUNTRYISO2': iso2, 'POSTCODE': None, 'MIN_WEIGHT': None,
            'MIN_VOLUME': None, 'MIN_PARCEL': None,
            'USER_DEF_TYPE_2': None,
            'USER_DEF_TYPE_4 (max 1,5m)': None, 'AWKWARD': None, 'RATE_EXTRA': 0}


def build_combined_weight_rows(c0, bands, max_parcel, service_level,
                               max_ew=None, postcode=None, user_def_type_2=None,
                               min_parcel=1):
    """Combined-weight pricing: the band rate is the freight for the WHOLE
    shipment, looked up ONCE on the total payweight — never rate * parcel_count.

    Carriers/services billed on consolidated weight (DHL standard, UPS DE/NL/GB
    EXPRESS SAVER, UPS DE STDM) must use this instead of the per-parcel
    `rate * mp` model, which overprices multi-parcel shipments.

    `bands` is the output of collapse_same_rate_tiers(tiers) WITHOUT a weight
    cap, i.e. total-payweight bands as (band_top, rate, per_kg) tuples.

    EACH_WEIGHT is only a per-box cap chosen so that
    MAX_PARCEL * EACH_WEIGHT == the band's total-weight ceiling, keeping the
    existing writer/compute (MAX_WEIGHT = MAX_PARCEL * EACH_WEIGHT,
    MAX_VOLUME = MAX_WEIGHT / divisor) consistent.

    `max_ew` caps the per-box weight to the carrier's physical single-parcel
    maximum: a band ceiling is only reachable with enough parcels
    (band_top / mp <= max_ew). This prevents nonsensical rows like one parcel of
    250 kg and keeps the matrix from exploding, while still covering heavy
    multi-parcel shipments. EACH_WEIGHT is NOT snapped to the integer grid, so
    no reachable band/parcel combination is dropped.
    """
    rows = []
    for band_top, rate, per_kg in bands:
        if per_kg:                       # skip the open-ended "over X / kg" tail
            continue
        for mp in range(min_parcel, max_parcel + 1):
            each = band_top / mp
            if max_ew is not None and each > max_ew + 1e-9:
                continue                 # band unreachable with this few parcels
            row = {**c0, 'SERVICE_LEVEL': service_level,
                   'MAX_PARCEL': mp,
                   'EACH_WEIGHT': round(each, 6),            # cap; mp*each = band_top
                   'RATE_BASE': round(rate, 4)}              # ONE lookup, no * mp
            if postcode is not None:
                row['POSTCODE'] = postcode
            if user_def_type_2 is not None:
                row['USER_DEF_TYPE_2'] = user_def_type_2
            rows.append(row)
    return rows


def build_rows_upde(rate_data, country_cfg):
    rows   = []
    max_p  = country_cfg['max_parcel_count']
    max_ew = country_cfg['max_each_weight_kg']
    c0     = _common(country_cfg['site_id'], country_cfg['client_id'],
                     'UPDE', country_cfg['iso2'])

    # STANDARD — combined weight. A single parcel (mp=1) is priced on the STDS
    # table (lookup on the parcel's own weight, which IS the combined weight);
    # multi-parcel shipments (mp>=2) on the STDM total-weight table. Both do ONE
    # rate lookup on the band ceiling, never rate * parcel_count.
    for pc, tiers in _upde_service_buckets(rate_data, 'STDS', country_cfg):
        bands = collapse_same_rate_tiers(tiers)
        rows += build_combined_weight_rows(
            c0, bands, max_parcel=1, service_level='STANDARD',
            max_ew=max_ew, postcode=pc, user_def_type_2='single')

    for pc, tiers in _upde_service_buckets(rate_data, 'STDM', country_cfg):
        bands = collapse_same_rate_tiers(tiers)
        rows += build_combined_weight_rows(
            c0, bands, max_p, service_level='STANDARD',
            max_ew=max_ew, postcode=pc, user_def_type_2='multi', min_parcel=2)

    flat = rate_data.get('EXPSAVER_7R9W62')
    if flat is not None:
        for mp in range(1, max_p + 1):
            rows.append({**c0, 'SERVICE_LEVEL': 'EXPRESS SAVER 7R9W62',
                         'MAX_PARCEL': mp, 'EACH_WEIGHT': max_ew,
                         'RATE_BASE': round(flat * mp, 4)})

    for pc, tiers in _upde_service_buckets(rate_data, 'EXPRESS_SAVER', country_cfg):
        # EXPRESS SAVER is billed on the TOTAL shipment payweight (one lookup),
        # not per parcel — use the full total-payweight bands (no max_ew cap).
        bands = collapse_same_rate_tiers(tiers)
        rows += build_combined_weight_rows(c0, bands, max_p, 'EXPRESS SAVER',
                                           max_ew=max_ew, postcode=pc)

    # ---- WorldEase (WEA): flat per-country rate (CH, NO) ----
    wea = rate_data.get('WEA')
    if wea is not None:
        for mp in range(1, max_p + 1):
            rows.append({**c0, 'SERVICE_LEVEL': 'WORLDEASE',
                         'MAX_PARCEL': mp, 'EACH_WEIGHT': max_ew,
                         'RATE_BASE': round(wea * mp, 4)})

    return rows


def build_rows_dhl(rate_data, country_cfg):
    rows  = []
    max_p = country_cfg['max_parcel_count']
    max_ew= country_cfg['max_each_weight_kg']
    c0    = _common(country_cfg['site_id'], country_cfg['client_id'],
                    'DHL-ROS', country_cfg['iso2'])

    # ---- BNL pricing: 1st parcel + each-additional, no weight tiers (BE/LU/NL) ----
    bnl = rate_data.get('bnl')
    if bnl:
        first = bnl.get('first')
        after = bnl.get('after', first)
        if first is not None:
            for mp in range(1, max_p + 1):
                total = first + (mp - 1) * after
                rows.append({**c0, 'SERVICE_LEVEL': 'STANDARD',
                             'MAX_PARCEL': mp, 'EACH_WEIGHT': max_ew,
                             'RATE_BASE': round(total, 4)})
        return rows

    # DHL "Other countries" is a TOTAL-payweight table (one lookup per shipment),
    # not a per-parcel table — use the full bands (no max_ew cap) and price once.
    bands = collapse_same_rate_tiers(rate_data.get('STANDARD', []))
    rows += build_combined_weight_rows(c0, bands, max_p, 'STANDARD', max_ew=max_ew)
    return rows


def build_rows_dpd(rate_data, country_cfg):
    rows  = []
    max_p = country_cfg['max_parcel_count']
    c0    = _common(country_cfg['site_id'], country_cfg['client_id'],
                    'DPD', country_cfg['iso2'])
    for mp in range(1, max_p + 1):
        if 'klein' in rate_data:
            rows.append({**c0, 'SERVICE_LEVEL': 'PARCEL', 'MAX_PARCEL': mp,
                         'EACH_WEIGHT': 3.0,
                         'RATE_BASE': round(rate_data['klein'] * mp, 4)})
        if 'groot' in rate_data:
            rows.append({**c0, 'SERVICE_LEVEL': 'PARCEL', 'MAX_PARCEL': mp,
                         'EACH_WEIGHT': 31.5,
                         'RATE_BASE': round(rate_data['groot'] * mp, 4)})
    return rows


def build_rows_upsnl(rate_data, country_cfg):
    # EXPRESS SAVER is billed on the TOTAL shipment payweight (one lookup),
    # so every zone uses combined-weight rows (no per-parcel * mp, no max_ew cap).
    rows   = []
    max_p  = country_cfg['max_parcel_count']
    max_ew = country_cfg['max_each_weight_kg']
    pc_min, pc_max = country_cfg['postcode_prefix_range']
    c0     = _common(country_cfg['site_id'], country_cfg['client_id'],
                     'UPSNL', country_cfg['iso2'])

    zones = rate_data.get('zones', [])
    bands_by_zone = {z: collapse_same_rate_tiers(t)
                     for z, t in rate_data.get('rates_by_zone', {}).items()}

    def emit(zone, pc):
        return build_combined_weight_rows(
            c0, bands_by_zone.get(zone, []), max_p, 'EXPRESS SAVER',
            max_ew=max_ew, postcode=pc)

    # No zone table at all → no postcode, use first available zone
    if not zones:
        zone = next(iter(bands_by_zone), None)
        return emit(zone, None) if zone is not None else rows

    # Has pc_from/pc_to ranges → map each prefix to a zone
    has_pc_ranges = any('pc_from' in z and 'pc_to' in z for z in zones)
    if not has_pc_ranges:
        # Single-zone entry without postcode ranges → no postcode column
        return emit(zones[0].get('zone'), None)

    # Worst-zone fallback for prefixes not covered by the table
    def _zone_max_rate(zid):
        return max((r for _, r, _ in bands_by_zone.get(zid, [])), default=0)
    all_zone_ids  = [z['zone'] for z in zones if 'zone' in z]
    fallback_zone = max(all_zone_ids, key=_zone_max_rate) if all_zone_ids else None

    prefix_to_zone = {}
    for pc_prefix in range(pc_min, pc_max + 1):
        pc_full = pc_prefix * 1000
        matched = next((z['zone'] for z in zones
                        if z.get('pc_from', 0) <= pc_full <= z.get('pc_to', 99999)
                        and 'zone' in z), fallback_zone)
        if matched is not None:
            prefix_to_zone[pc_prefix] = matched

    unique_zones = set(prefix_to_zone.values())
    if len(unique_zones) <= 1:
        rows += emit(next(iter(unique_zones), fallback_zone), None)
    else:
        for pc_prefix, zone in prefix_to_zone.items():
            rows += emit(zone, pc_prefix)
    return rows


def build_rows_postnord(rate_data, country_cfg):
    """
    Handles two PostNord formats:
      Format A — flat_rates dict  {'B2B': 12.2, 'HOME': 12.9, 'PUDO': 10.25}
                 → one row per (service, parcel_count) at max_each_weight
      Format B — weight-tier list {'STANDARD': [{from,to,rate,per_kg},...]}
                 → one row per (weight_band, parcel_count)
    """
    rows   = []
    max_p  = country_cfg['max_parcel_count']
    max_ew = country_cfg['max_each_weight_kg']
    c0     = _common(country_cfg['site_id'], country_cfg['client_id'],
                     'POSTNORD', country_cfg['iso2'])

    flat_rates = rate_data.get('flat_rates', {})
    if flat_rates:
        for svc_name, rate in flat_rates.items():
            for mp in range(1, max_p + 1):
                rows.append({**c0, 'SERVICE_LEVEL': svc_name,
                             'MAX_PARCEL': mp, 'EACH_WEIGHT': max_ew,
                             'RATE_BASE': round(rate * mp, 4)})
    else:
        for each_w, rate, per_kg in collapse_same_rate_tiers(
                rate_data.get('STANDARD', []), max_ew):
            if each_w > max_ew or per_kg:
                continue
            for mp in range(1, max_p + 1):
                rows.append({**c0, 'SERVICE_LEVEL': 'STANDARD',
                             'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                             'RATE_BASE': round(rate * mp, 4)})
    return rows


def build_rows_upsgb(rate_data, country_cfg):
    """UPS GB (UK domestic): STDS (single, per-parcel), STDM (combined weight),
    EXPS (express). Single rate column, no postcode zones. Linehaul applied
    via carrier_defaults."""
    rows   = []
    max_p  = country_cfg['max_parcel_count']
    max_ew = country_cfg['max_each_weight_kg']
    c0     = _common(country_cfg['site_id'], country_cfg['client_id'],
                     'UPSGB', country_cfg['iso2'])

    # STANDARD — combined weight: mp=1 priced on STDS (single), mp>=2 on STDM
    # (multi, total-weight). One rate lookup per band, never rate * parcel_count.
    bands_stds = collapse_same_rate_tiers(rate_data.get('STDS', []))
    rows += build_combined_weight_rows(
        c0, bands_stds, max_parcel=1, service_level='STANDARD',
        max_ew=max_ew, user_def_type_2='single')

    bands_stdm = collapse_same_rate_tiers(rate_data.get('STDM', []))
    rows += build_combined_weight_rows(
        c0, bands_stdm, max_p, service_level='STANDARD',
        max_ew=max_ew, user_def_type_2='multi', min_parcel=2)

    # EXPS — express saver: billed on TOTAL shipment payweight (one lookup).
    bands = collapse_same_rate_tiers(rate_data.get('EXPS', []))
    rows += build_combined_weight_rows(c0, bands, max_p, 'EXPRESS SAVER', max_ew=max_ew)
    return rows


CARRIER_BUILDERS = {
    'UPDE':     build_rows_upde,
    'DHL-ROS':  build_rows_dhl,
    'DPD':      build_rows_dpd,
    'UPSNL':    build_rows_upsnl,
    'POSTNORD': build_rows_postnord,
    'UPSGB':    build_rows_upsgb,
}


def build_extended_matrix(parsed, country_cfg):
    all_rows = []
    for carrier in country_cfg['carriers']:
        builder = CARRIER_BUILDERS.get(carrier)
        if builder is None:
            log.warning("No builder for carrier '%s'", carrier)
            continue
        data = parsed.get(carrier, {})
        if not data:
            log.warning("No rate data for '%s', skipping", carrier)
            continue
        carrier_rows = builder(data, country_cfg)
        log.info('  %s: %d rows', carrier, len(carrier_rows))
        all_rows.extend(carrier_rows)
    return pd.DataFrame(all_rows)


# ==============================================================================
# 6. NUMERIC PRE-COMPUTATION
# ==============================================================================

def compute_numeric_totals(df, carrier_defaults=None):
    cd  = carrier_defaults or CARRIER_DEFAULTS
    df  = df.copy()
    if df.empty:
        return df

    df['MAX_WEIGHT'] = df['MAX_PARCEL'] * df['EACH_WEIGHT']

    df['FUEL'] = df.apply(
        lambda r: cd[r['CARRIER_ID']]['fuel_pct'] * r['RATE_BASE'], axis=1
    ).round(4)
    df['MAUT'] = df.apply(
        lambda r: cd[r['CARRIER_ID']]['maut_pct'] * r['RATE_BASE'], axis=1
    ).round(4)
    df['Linehaul UPSDE'] = pd.to_numeric(df.apply(
        lambda r: (cd[r['CARRIER_ID']]['linehaul_per_parcel'] * r['MAX_PARCEL']
                   if cd[r['CARRIER_ID']]['linehaul_per_parcel'] > 0
                   and r['MAX_PARCEL'] is not None
                   and not pd.isna(r['MAX_PARCEL']) else None),
        axis=1,
    ), errors='coerce').round(4)

    vdiv = lambda r: cd[r['CARRIER_ID']]['volume_divisor']
    df['MAX_VOLUME']  = df.apply(lambda r: r['MAX_WEIGHT'] / vdiv(r), axis=1)
    df['EACH_VOLUME'] = df.apply(lambda r: r['EACH_WEIGHT'] / vdiv(r), axis=1)

    df['TOTAL_PRICE'] = (
        df['RATE_BASE'] + df['RATE_EXTRA'].fillna(0) + df['FUEL'] + df['MAUT']
        + df['Linehaul UPSDE'].fillna(0)
    ).round(4)
    return df


# ==============================================================================
# 7. EXCEL WRITER
# ==============================================================================

COLUMN_ORDER = [
    'SITE_ID', 'CLIENT_ID', 'CARRIER_ID', 'SERVICE_LEVEL', 'COUNTRYISO2',
    'POSTCODE', 'MIN_WEIGHT', 'MAX_WEIGHT', 'MIN_VOLUME', 'MAX_VOLUME',
    'MIN_PARCEL', 'MAX_PARCEL', 'EACH_WEIGHT', 'EACH_VOLUME',
    'USER_DEF_TYPE_2',
    'USER_DEF_TYPE_4 (max 1,5m)', 'AWKWARD', 'RATE_BASE', 'RATE_EXTRA',
    'FUEL', 'MAUT', 'Linehaul UPSDE', 'TOTAL_PRICE',
]
COL_LETTER = {name: openpyxl.utils.get_column_letter(i + 1)
              for i, name in enumerate(COLUMN_ORDER)}


def _build_formulas_for_row(row_dict, excel_row, carrier_defaults=None):
    cd  = carrier_defaults or CARRIER_DEFAULTS
    L   = COL_LETTER
    cfg = cd[row_dict['CARRIER_ID']]
    f   = {}
    # Overflow buckets leave MAX_PARCEL / EACH_WEIGHT blank — skip the grid
    # formulas for them so we don't emit "=*" ; their values stay literal (None).
    has_grid = (row_dict.get('MAX_PARCEL') is not None
                and row_dict.get('EACH_WEIGHT') is not None)
    if has_grid:
        f['MAX_WEIGHT']  = f"={L['MAX_PARCEL']}{excel_row}*{L['EACH_WEIGHT']}{excel_row}"
        f['MAX_VOLUME']  = f"={L['MAX_WEIGHT']}{excel_row}/{cfg['volume_divisor']}"
        f['EACH_VOLUME'] = f"={L['EACH_WEIGHT']}{excel_row}/{cfg['volume_divisor']}"
    if cfg.get('fuel_variables_ref'):
        ref = cfg['fuel_variables_ref']
        f['FUEL'] = f"=Variables!${ref[0]}${ref[1:]}*{L['RATE_BASE']}{excel_row}"
    if cfg.get('maut_variables_ref'):
        ref = cfg['maut_variables_ref']
        f['MAUT'] = f"=Variables!${ref[0]}${ref[1:]}*{L['RATE_BASE']}{excel_row}"
    lh = L['Linehaul UPSDE']
    f['TOTAL_PRICE'] = (
        f"={L['RATE_BASE']}{excel_row}+{L['RATE_EXTRA']}{excel_row}"
        f"+{L['FUEL']}{excel_row}+{L['MAUT']}{excel_row}"
        f'+{lh}{excel_row}'
    )
    return f


def write_matrix_excel(df, output_path, country_cfg,
                       carrier_defaults=None, variables_layout=None):
    vl  = variables_layout or VARIABLES_LAYOUT
    wb  = Workbook()
    ws  = wb.active
    ws.title = f"{country_cfg['iso2']} Matrix"
    for ci, col in enumerate(COLUMN_ORDER, 1):
        ws.cell(1, ci, col)
    df_sorted = df.sort_values('TOTAL_PRICE', kind='stable').reset_index(drop=True)
    bucket_fill = PatternFill('solid', fgColor='FFF2CC')   # soft amber = catch-all bucket
    for ri, row_dict in enumerate(df_sorted.to_dict('records'), start=2):
        is_bucket = bool(row_dict.get('_is_bucket'))
        # A sentinel catch-all (bucket with no rate components) keeps its literal
        # TOTAL_PRICE; building a formula would sum blanks to 0.
        sentinel = is_bucket and pd.isna(row_dict.get('RATE_BASE'))
        formulas = {} if sentinel else _build_formulas_for_row(row_dict, ri, carrier_defaults)
        for ci, col in enumerate(COLUMN_ORDER, 1):
            if col in formulas:
                cell = ws.cell(ri, ci, formulas[col])
            else:
                val = row_dict.get(col)
                cell = ws.cell(ri, ci, None if pd.isna(val) else val)
            if is_bucket:
                cell.fill = bucket_fill
    vs = wb.create_sheet('Variables')
    for ri, (name, val) in enumerate(vl, 1):
        vs.cell(ri, 1, name)
        vs.cell(ri, 2, val)
    wb.save(output_path)
    log.info('wrote %s (%d rows)', output_path, len(df_sorted))


# ==============================================================================
# 8. FIRST-PASS OPTIMIZER (per carrier/service)
# ==============================================================================

def optimize_matrix(df):
    df = df.reset_index(drop=True)
    if df.empty:
        return df
    # str key so numeric / UK-alpha / blank postcodes share a group cleanly.
    pk = df['POSTCODE'].fillna('').astype(str).values
    car = df['CARRIER_ID'].values
    svc = df['SERVICE_LEVEL'].values
    w = df['MAX_WEIGHT'].values.astype(float)
    p = df['MAX_PARCEL'].values.astype(float)
    e = df['EACH_WEIGHT'].values.astype(float)
    price = df['TOTAL_PRICE'].values.astype(float)
    groups = {}
    for idx in range(len(df)):
        groups.setdefault((car[idx], svc[idx], pk[idx]), []).append(idx)
    drop = set()
    for idxs in groups.values():
        idxs = sorted(idxs, key=lambda k: price[k])     # cheapest first
        for a in range(1, len(idxs)):
            i = idxs[a]
            ear = np.asarray(idxs[:a])
            if ((w[ear] >= w[i]) & (p[ear] >= p[i]) & (e[ear] >= e[i])).any():
                drop.add(i)
    keep = [i for i in range(len(df)) if i not in drop]
    log.info('first-pass: removed %d dominated rows', len(drop))
    return df.iloc[keep].reset_index(drop=True)


# ==============================================================================
# 9. GLOBAL OPTIMIZER (cross-carrier)
# ==============================================================================

_RELATIVE_REF = re.compile(r'(\$?[A-Z]+)(\$?)(\d+)')


def _update_formula(formula, old_row, new_row):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    def repl(m):
        col, dollar, row = m.group(1), m.group(2), m.group(3)
        if dollar == '$':
            return m.group(0)
        return f'{col}{new_row}' if int(row) == old_row else m.group(0)
    return _RELATIVE_REF.sub(repl, formula)


def _write_filtered_excel(input_path, output_path, keep_indices):
    shutil.copy(input_path, output_path)
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    ncols = ws.max_column
    keep_excel = {idx + 2 for idx in keep_indices}
    kept = [(r, [ws.cell(r, c).value for c in range(1, ncols + 1)])
            for r in range(2, ws.max_row + 1) if r in keep_excel]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for new_row, (old_row, vals) in enumerate(kept, start=2):
        for c, val in enumerate(vals, start=1):
            if isinstance(val, str) and val.startswith('='):
                val = _update_formula(val, old_row, new_row)
            ws.cell(new_row, c, val)
    wb.save(output_path)
    return len(kept)


def _ensure_numeric(df, input_path):
    df = df.copy()
    if df['MAX_WEIGHT'].isna().any():
        df['MAX_WEIGHT'] = df['MAX_PARCEL'] * df['EACH_WEIGHT']
    variables = {}
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        if 'Variables' in wb.sheetnames:
            vs = wb['Variables']
            for r in range(1, vs.max_row + 1):
                name = vs.cell(r, 1).value
                val  = vs.cell(r, 2).value
                if isinstance(name, str) and isinstance(val, (int, float)):
                    variables[name.strip()] = float(val)
    except Exception as e:
        log.warning('Could not read Variables sheet: %s', e)
    fuel_map = {'UPDE': 'FUEL UPSDE', 'DHL-ROS': 'FUEL DHL',
                'DPD': 'FUEL DPD', 'UPSNL': 'FUEL UPSNL',
                'POSTNORD': 'FUEL POSTNORD'}
    maut_map = {'DPD': 'MAUT DPD', 'DHL-ROS': 'MAUT DHL'}
    pct = lambda carrier, mapping: variables.get(mapping.get(carrier, ''), 0.0)
    if df['FUEL'].isna().any():
        df['FUEL'] = df.apply(
            lambda r: pct(r['CARRIER_ID'], fuel_map) * (r['RATE_BASE'] or 0), axis=1)
    if df['MAUT'].isna().any():
        df['MAUT'] = df.apply(
            lambda r: pct(r['CARRIER_ID'], maut_map) * (r['RATE_BASE'] or 0), axis=1)
    if df['TOTAL_PRICE'].isna().any():
        df['TOTAL_PRICE'] = (df['RATE_BASE'].fillna(0) + df['RATE_EXTRA'].fillna(0)
                             + df['FUEL'].fillna(0) + df['MAUT'].fillna(0)
                             + df['Linehaul UPSDE'].fillna(0))
    return df


def optimize_globally(input_path, output_path):
    df = pd.read_excel(input_path, sheet_name=0)
    df = _ensure_numeric(df, input_path)
    df_s = df.sort_values('TOTAL_PRICE', kind='stable').reset_index()
    df_s = df_s.rename(columns={'index': '_orig'})
    w  = df_s['MAX_WEIGHT'].values.astype(float)
    p  = df_s['MAX_PARCEL'].values.astype(float)
    e  = df_s['EACH_WEIGHT'].values.astype(float)
    pc = df_s['POSTCODE'].values                 # object: int / str (UK) / None
    pc_blank = pd.isna(pc)
    orig   = df_s['_orig'].values
    n = len(df_s)
    dominated = set()
    blank_idx = np.where(pc_blank)[0]

    def _mark(idxs, skip_blank):
        idxs = np.asarray(idxs)
        for a in range(1, len(idxs)):
            i = idxs[a]
            if skip_blank and pc_blank[i]:
                continue
            ear = idxs[:a]
            if ((w[ear] >= w[i]) & (p[ear] >= p[i]) & (e[ear] >= e[i])).any():
                dominated.add(int(orig[i]))

    _mark(blank_idx, skip_blank=False)
    groups = {}
    for j in range(n):
        if not pc_blank[j]:
            groups.setdefault(pc[j], []).append(j)
    blank_list = blank_idx.tolist()
    for idxs in groups.values():
        _mark(sorted(blank_list + idxs), skip_blank=True)
    keep = set(df.index) - dominated
    n = _write_filtered_excel(input_path, output_path, keep)
    log.info('global optimizer: %d → %d rows (%d removed)', len(df), n, len(dominated))
    return {'input_rows': len(df), 'removed': len(dominated),
            'output_rows': n, 'output_path': str(output_path)}


# ==============================================================================
# 9b. EXCEPTIONS / BUCKETS  (add-on)
# ==============================================================================
#
# CargoWrite matches an order by scanning the matrix top-to-bottom (cheapest
# first) and taking the first row whose constraints all fit. A constraint such
# as USER_DEF_TYPE_4 (a max dimension) means "this row only matches parcels at
# or under this size". Orders that exceed it must still match *something*, so
# every constrained row needs a cheaper-to-build "bucket" twin lower down that
# drops the limit, flags the row for oversight, and adds a surcharge.
#
# An exception rule is a plain dict. Two modes:
#
# mode='stamp' (default) — for INERT constraint columns (size flags). Stamps a
# normal limit on the cheap base rows and appends a bucket TWIN that drops the
# limit, flags it, and adds a surcharge. Orders bigger than the limit fall
# through to the twin.
#   {
#     'enabled':        True,
#     'label':          'Oversize (max 1.5m)',
#     'carriers':       ['UPDE'],      # scope; [] = all carriers
#     'countries':      [],            # scope; [] = all countries
#     'service_levels': [],            # scope; [] = all services
#     'mode':           'stamp',       # (default)
#     'constraint_col': 'USER_DEF_TYPE_4 (max 1,5m)',
#     'normal_value':   1.5,           # stamped on the cheap base rows
#     'bucket_value':   None,          # value on the bucket twin (None = catch-all)
#     'flag_col':       'AWKWARD',     # column flagged on the bucket twin
#     'flag_value':     'y',
#     'surcharge':      6.0,           # euros
#     'surcharge_mode': 'per_parcel',  # 'per_parcel' (× MAX_PARCEL) or 'flat'
#   }
#
# mode='threshold' — for COMPUTATIONAL constraint columns (e.g. EACH_WEIGHT,
# which feeds MAX_WEIGHT/volumes and must NOT be overwritten). Does NOT stamp or
# twin: it surcharges, in place, every in-scope BASE row whose constraint value
# is >= threshold, flags it, and marks it as a bucket (amber). No twin is made
# because an identical-constraint twin would never be matched (the matcher always
# takes the cheaper original). Use for per-parcel weight surcharges such as
# "DHL parcel >= 20 kg costs €4.89 per parcel more".
#   {
#     'enabled':        True,
#     'label':          'Heavy parcel DHL >=20kg',
#     'carriers':       ['DHL-ROS'],   # scope
#     'countries':      [],            # scope
#     'mode':           'threshold',
#     'constraint_col': 'EACH_WEIGHT', # per-box cap (kg) — NOT overwritten
#     'threshold':      20.0,          # surcharge rows whose cap can hold a >=20kg box
#     'flag_col':       'AWKWARD',
#     'flag_value':     'y',
#     'surcharge':      4.89,          # euros
#     'surcharge_mode': 'per_parcel',  # 'per_parcel' (× MAX_PARCEL) or 'flat'
#   }
#
# NOTE: a heavy-parcel threshold and an oversize stamp do not stack on the same
# shipment (the oversize twin does not inherit the heavy surcharge). This affects
# only the rare parcel that is BOTH over the size limit AND >= the weight
# threshold; the common heavy-but-normal-size parcel is surcharged correctly.
# Because EACH_WEIGHT is a derived per-box cap (band_top / parcel_count), a row
# whose cap is >= the threshold may also match an all-light shipment that happens
# to fall in that band — such a shipment is over-charged, never under-charged
# (the safe direction, consistent with the bucket philosophy).
#
# FUTURE BUCKET IDEAS (designed for, not yet implemented):
#   • parcel-count overflow  — a row with MAX_PARCEL blank to catch >max parcels
#   • weight overflow        — a row with MAX_WEIGHT blank for over-grid weights
#   • postcode catch-all     — a blank-POSTCODE row when zone rows miss a prefix
#   apply_exceptions already leaves room for these: add a rule whose
#   constraint_col is MAX_PARCEL / MAX_WEIGHT / POSTCODE and bucket_value=None.


def _recompute_total(df):
    df['TOTAL_PRICE'] = (
        df['RATE_BASE'].fillna(0) + df['RATE_EXTRA'].fillna(0)
        + df['FUEL'].fillna(0) + df['MAUT'].fillna(0)
        + df['Linehaul UPSDE'].fillna(0)
    ).round(4)
    return df


def apply_exceptions(df, rules):
    """Stamp normal limits on in-scope base rows and append bucket twins.

    Returns a new DataFrame with an extra boolean column '_is_bucket'
    (write_matrix_excel uses it to colour the bucket rows). The base matrix is
    expected to already carry numeric FUEL / MAUT / Linehaul columns.
    """
    if not rules:
        return df

    df = df.copy().reset_index(drop=True)
    if '_is_bucket' not in df.columns:
        df['_is_bucket'] = False
    # RATE_EXTRA may arrive as int64 (initialised to 0); coerce to float so the
    # in-place threshold surcharge assignment cannot raise a dtype error.
    if 'RATE_EXTRA' in df.columns:
        df['RATE_EXTRA'] = pd.to_numeric(df['RATE_EXTRA'], errors='coerce').astype('float64')

    new_buckets = []
    touched = False
    for rule in rules:
        if not rule.get('enabled', True):
            continue
        ccol = rule['constraint_col']
        if ccol not in df.columns:
            log.warning("exception rule skipped — column '%s' not in matrix", ccol)
            continue

        # in-scope BASE rows (buckets are never re-processed)
        scope = (~df['_is_bucket'])
        if rule.get('carriers'):
            scope &= df['CARRIER_ID'].isin(rule['carriers'])
        if rule.get('countries'):
            scope &= df['COUNTRYISO2'].isin(rule['countries'])
        if rule.get('service_levels'):
            scope &= df['SERVICE_LEVEL'].isin(rule['service_levels'])
        if not scope.any():
            continue

        sur        = float(rule.get('surcharge', 0) or 0)
        per_parcel = rule.get('surcharge_mode', 'per_parcel') == 'per_parcel'
        mode       = rule.get('mode', 'stamp')

        if mode == 'threshold':
            # Per-parcel weight/size THRESHOLD on a COMPUTATIONAL column.
            # Do NOT overwrite ccol (it feeds MAX_WEIGHT/volumes). Surcharge the
            # rows whose per-box cap can hold a parcel at/over the threshold,
            # flag them, and mark them as buckets (amber). No twins.
            try:
                thr = float(rule['threshold'])
            except (TypeError, ValueError, KeyError):
                log.warning("threshold rule '%s' skipped — no valid 'threshold'",
                            rule.get('label', ccol))
                continue
            col_num = pd.to_numeric(df[ccol], errors='coerce')
            hit = scope & (col_num >= thr - 1e-9)
            if not hit.any():
                continue
            if per_parcel:
                add = sur * df.loc[hit, 'MAX_PARCEL'].fillna(1)
            else:
                add = sur
            df.loc[hit, 'RATE_EXTRA'] = df.loc[hit, 'RATE_EXTRA'].fillna(0) + add
            if rule.get('flag_col'):
                df.loc[hit, rule['flag_col']] = rule.get('flag_value', 'y')
            df.loc[hit, '_is_bucket'] = True
            touched = True
            continue

        # ---- mode == 'stamp' (original behaviour) ----
        # 1) stamp the normal limit onto the cheap base rows
        df.loc[scope, ccol] = rule['normal_value']

        # 2) build the bucket twins (limit removed, flagged, surcharged)
        twins = df[scope].copy()
        twins[ccol] = rule.get('bucket_value', None)
        if rule.get('flag_col'):
            twins[rule['flag_col']] = rule.get('flag_value', 'y')
        if per_parcel:
            twins['RATE_EXTRA'] = twins['RATE_EXTRA'].fillna(0) + sur * twins['MAX_PARCEL']
        else:
            twins['RATE_EXTRA'] = twins['RATE_EXTRA'].fillna(0) + sur
        twins['_is_bucket'] = True
        new_buckets.append(twins)
        touched = True

    if new_buckets:
        df = pd.concat([df] + new_buckets, ignore_index=True)
    if touched:
        df = _recompute_total(df)
    return df


def add_overflow_buckets(df, rules, carrier_defaults=None, country_cfg=None):
    """Append combined parcel+weight overflow buckets (matches the corrected
    example's structure). For each in-scope (carrier, service, country) and each
    parcel count n in 1..grid_max_parcels, add a catch-all row:

        MIN_PARCEL = n,            MAX_PARCEL = blank   (catches >= n parcels)
        MIN_WEIGHT = n*grid_max,   MAX_WEIGHT = blank   (catches over-grid weight)
        EACH_WEIGHT = blank
        RATE_BASE  = overflow_rate * n       (overflow_rate is manager-supplied)
        RATE_EXTRA = surcharge * n
        flag       = AWKWARD = 'Y'

    The overflow_rate is NOT inferred from the grid (the example used a hand-set
    heavy/per-kg rate); it must be provided in the rule. Rows are flagged so ops
    can review every overflow shipment.
    """
    cd = carrier_defaults or CARRIER_DEFAULTS
    if not rules:
        return df
    df = df.copy().reset_index(drop=True)
    if '_is_bucket' not in df.columns:
        df['_is_bucket'] = False

    new = []
    for rule in rules:
        if not rule.get('enabled', True):
            continue
        try:
            rate = float(rule['overflow_rate'])
        except (TypeError, ValueError, KeyError):
            log.warning('overflow rule skipped — no valid overflow_rate'); continue
        sur      = float(rule.get('surcharge', 0) or 0)
        flag_col = rule.get('flag_col', 'AWKWARD')
        flag_val = rule.get('flag_value', 'Y')

        base = df[~df['_is_bucket']]
        if rule.get('carriers'):
            base = base[base['CARRIER_ID'].isin(rule['carriers'])]
        if rule.get('countries'):
            base = base[base['COUNTRYISO2'].isin(rule['countries'])]
        if base.empty:
            continue

        for (carrier, svc, country), grp in base.groupby(
                ['CARRIER_ID', 'SERVICE_LEVEL', 'COUNTRYISO2']):
            # Grid ceilings come from the country config (the TRUE limits), not
            # the post-optimization frame where weight bands have collapsed.
            if country_cfg:
                max_p    = int(country_cfg['max_parcel_count'])
                grid_max = float(country_cfg['max_each_weight_kg'])
            else:
                max_p    = int(grp['MAX_PARCEL'].max())
                grid_max = float(grp['EACH_WEIGHT'].max())
            site, client = grp.iloc[0]['SITE_ID'], grp.iloc[0]['CLIENT_ID']
            for n in range(1, max_p + 1):
                row = _common(site, client, carrier, country)
                rb  = round(rate * n, 4)
                fuel = round(cd[carrier]['fuel_pct'] * rb, 4)
                maut = round(cd[carrier]['maut_pct'] * rb, 4)
                lh_pp = cd[carrier]['linehaul_per_parcel']
                lh   = round(lh_pp * n, 4) if lh_pp > 0 else None
                row.update({
                    'SERVICE_LEVEL': svc,
                    'MIN_PARCEL': n,       'MAX_PARCEL': None,
                    'MIN_WEIGHT': round(n * grid_max, 4), 'MAX_WEIGHT': None,
                    'EACH_WEIGHT': None,   'MAX_VOLUME': None, 'EACH_VOLUME': None,
                    'RATE_BASE': rb,       'RATE_EXTRA': round(sur * n, 4),
                    'FUEL': fuel,          'MAUT': maut, 'Linehaul UPSDE': lh,
                    flag_col: flag_val,    '_is_bucket': True,
                })
                row['TOTAL_PRICE'] = round(rb + sur * n + fuel + maut + (lh or 0), 4)
                new.append(row)

    if new:
        df = pd.concat([df, pd.DataFrame(new)], ignore_index=True)
    return df


def add_postcode_catchall(df, rules, carrier_defaults=None):
    """For zoned carriers (rows carrying a specific POSTCODE prefix), append a
    POSTCODE=blank fallback at the worst (most expensive) zone's rate, flagged,
    so a prefix not present in any zone still matches something."""
    if not rules:
        return df
    df = df.copy().reset_index(drop=True)
    if '_is_bucket' not in df.columns:
        df['_is_bucket'] = False

    new = []
    for rule in rules:
        if not rule.get('enabled', True):
            continue
        sur      = float(rule.get('surcharge', 0) or 0)
        flag_col = rule.get('flag_col', 'AWKWARD')
        flag_val = rule.get('flag_value', 'Y')

        base = df[(~df['_is_bucket']) & (df['POSTCODE'].notna())]
        if rule.get('carriers'):
            base = base[base['CARRIER_ID'].isin(rule['carriers'])]
        if rule.get('countries'):
            base = base[base['COUNTRYISO2'].isin(rule['countries'])]
        if base.empty:
            continue

        # worst-case row per (carrier, service, country, parcels, each-weight)
        keys = ['CARRIER_ID', 'SERVICE_LEVEL', 'COUNTRYISO2', 'MAX_PARCEL', 'EACH_WEIGHT']
        worst = base.loc[base.groupby(keys)['TOTAL_PRICE'].idxmax()].copy()
        worst['POSTCODE']   = None
        worst['RATE_EXTRA'] = worst['RATE_EXTRA'].fillna(0) + sur
        worst[flag_col]     = flag_val
        worst['_is_bucket'] = True
        new.append(worst)

    if new:
        df = pd.concat([df] + new, ignore_index=True)
        df = _recompute_total(df)
    return df


def optimize_globally_df(df):
    """Cross-carrier dominance on a DataFrame; returns the kept rows.
    Same logic as optimize_globally but stays in pandas (no Excel round-trip),
    so the exception/bucket step can run on the result before writing."""
    df = df.reset_index(drop=True)
    if df.empty:
        return df
    df_s = df.sort_values('TOTAL_PRICE', kind='stable').reset_index()
    df_s = df_s.rename(columns={'index': '_orig'})
    w  = df_s['MAX_WEIGHT'].values.astype(float)
    p  = df_s['MAX_PARCEL'].values.astype(float)
    e  = df_s['EACH_WEIGHT'].values.astype(float)
    pc = df_s['POSTCODE'].values                 # object: int / str (UK) / None
    pc_blank = pd.isna(pc)
    orig = df_s['_orig'].values
    n = len(df_s)
    dominated = set()

    # Postcode-partitioned dominance. A blank (country-wide) row can dominate
    # anything; a specific-postcode row can only be dominated by a blank row or
    # another row with the SAME postcode. Two different specific postcodes are
    # always incompatible, so comparing them is wasted work. Scanning blank rows
    # once, then each postcode group against (blank + itself), turns the whole-
    # matrix O(n^2) into the sum of small per-postcode scans — essential once GB
    # carries ~124 UK outward-code areas (was a multi-minute hang otherwise).
    blank_idx = np.where(pc_blank)[0]

    def _mark(idxs, skip_blank):
        idxs = np.asarray(idxs)
        for a in range(1, len(idxs)):
            i = idxs[a]
            if skip_blank and pc_blank[i]:
                continue
            ear = idxs[:a]                       # cheaper-or-equal candidates
            if ((w[ear] >= w[i]) & (p[ear] >= p[i]) & (e[ear] >= e[i])).any():
                dominated.add(int(orig[i]))

    _mark(blank_idx, skip_blank=False)           # blank rows: only blanks dominate
    groups = {}
    for j in range(n):
        if not pc_blank[j]:
            groups.setdefault(pc[j], []).append(j)
    blank_list = blank_idx.tolist()
    for idxs in groups.values():
        merged = sorted(blank_list + idxs)       # ascending price (df_s is sorted)
        _mark(merged, skip_blank=True)           # only mark the specific rows
    keep = [i for i in range(len(df)) if i not in dominated]
    return df.iloc[keep].reset_index(drop=True)


# ==============================================================================
# 9c. PARCEL POSTCODE EXPANSION
# ==============================================================================
#
# CargoWrite skips a parcel row that has no POSTCODE when the order being priced
# carries one — so every parcel row must ship with an explicit postcode for its
# country. Pallet (DHL-FENDER) rows already carry their per-zip postcode and are
# left untouched; parcel rows that already carry a specific postcode (zoned UPDE
# / UPSNL buckets) are also left as-is. Only the BLANK-postcode parcel rows are
# replicated, once per country postcode prefix.
#
# The postcode list is the country's set of prefixes from the DHL pallet file
# (the de-facto per-country list: numeric prefixes as text with leading zeros,
# e.g. '01','08','00'; Ireland uses named regions: DUBLIN, CORK, BT, …). Codes
# stay as text so '01' never collapses to 1.

def explode_parcel_postcodes(df, postcodes):
    """Replicate each blank-POSTCODE parcel row once per postcode prefix.

    Returns (df_out, n_base) where n_base is the number of blank parcel rows that
    were exploded. If `postcodes` is empty but blank parcel rows exist, returns
    the frame unchanged with n_base = -1 (caller should warn — the country has no
    postcode list, e.g. LI / SM, and its parcel rows will stay blank).
    """
    if df is None or df.empty:
        return df, 0
    codes = [str(p).strip() for p in (postcodes or []) if str(p).strip() != '']
    is_parcel = df['CARRIER_ID'].astype(str) != 'DHL-FENDER'
    pc = df['POSTCODE']
    blank = pc.isna() | (pc.astype(str).str.strip().isin(['', 'None', 'nan']))
    target = is_parcel & blank
    if not target.any():
        return df, 0
    if not codes:
        return df, -1
    keep = df[~target].copy()
    base = df[target].copy().reset_index(drop=True)
    rep = base.loc[base.index.repeat(len(codes))].copy()
    rep['POSTCODE'] = codes * len(base)
    out = pd.concat([keep, rep], ignore_index=True)
    return out, len(base)


# ==============================================================================
# 10. ORCHESTRATOR
# ==============================================================================

def run_pipeline(input_path, country, output_dir='.',
                 country_cfg=None, carrier_defaults=None, variables_layout=None,
                 exceptions=None, overflow_rules=None, postcode_rules=None,
                 pallet_max_band_kg=None):
    """
    Full pipeline for one country.

    Parameters
    ----------
    input_path       : path to the rate-card Excel
    country          : ISO-2 code, e.g. 'DE'
    output_dir       : directory where output files are written
    country_cfg      : dict — overrides COUNTRY_CONFIG[country] entirely
    carrier_defaults : dict — overrides module-level CARRIER_DEFAULTS
    variables_layout : list — overrides module-level VARIABLES_LAYOUT

    Returns
    -------
    dict with keys: extended, optimized, minimal,
                    rows_extended, rows_optimized, rows_minimal
    """
    country = country.upper()
    cfg = country_cfg or COUNTRY_CONFIG.get(country)
    if cfg is None:
        raise ValueError(f"No configuration for country '{country}'.")
    cd  = carrier_defaults or CARRIER_DEFAULTS
    vl  = variables_layout or VARIABLES_LAYOUT

    log.info('=== Pipeline for %s ===', country)
    parsed = parse_rate_cards(input_path)
    return run_pipeline_from_parsed(parsed, country, output_dir, cfg, cd, vl,
                                    exceptions, overflow_rules, postcode_rules,
                                    pallet_max_band_kg=pallet_max_band_kg)


# ──────────────────────────────────────────────────────────────────────────────
# Last-resort EUROCONNECT catch-all bucket
# One per country, appended at the very end of the generated matrix. It matches
# any pallet order up to the cap and is priced absurdly high, so CargoWrite never
# leaves an order unmatched — the sentinel price guarantees it is only chosen when
# nothing cheaper fits. Components are blank, so the writers render it as a literal
# price (not a formula).
EUROCONNECT_BUCKET_MAX_WEIGHT = 24000.0          # kg ceiling
EUROCONNECT_BUCKET_DENSITY    = 330.0            # kg/m³ → MAX_VOLUME = weight / density
EUROCONNECT_BUCKET_PRICE      = 999999.0         # sentinel last-resort price (EUR)


def append_euroconnect_buckets(df, site_id='NLMOE01', client_id='NLFENDER'):
    """Append one DHL-FENDER / EUROCONNECT catch-all row per country in `df`."""
    if df is None or df.empty:
        return df
    countries = list(dict.fromkeys(df['COUNTRYISO2'].dropna()))
    if not countries:
        return df
    if 'SITE_ID' in df.columns and df['SITE_ID'].notna().any():
        site_id = df['SITE_ID'].dropna().iloc[0]
    if 'CLIENT_ID' in df.columns and df['CLIENT_ID'].notna().any():
        client_id = df['CLIENT_ID'].dropna().iloc[0]
    max_vol = round(EUROCONNECT_BUCKET_MAX_WEIGHT / EUROCONNECT_BUCKET_DENSITY, 8)
    rows = []
    for iso in countries:
        row = {c: None for c in df.columns}
        row.update({
            'SITE_ID': site_id, 'CLIENT_ID': client_id,
            'CARRIER_ID': 'DHL-FENDER', 'SERVICE_LEVEL': 'EUROCONNECT',
            'COUNTRYISO2': iso,
            'MAX_WEIGHT': EUROCONNECT_BUCKET_MAX_WEIGHT,
            'MAX_VOLUME': max_vol,
            'TOTAL_PRICE': EUROCONNECT_BUCKET_PRICE,
        })
        if '_is_bucket' in df.columns:
            row['_is_bucket'] = True
        rows.append(row)
    bucket_df = pd.DataFrame(rows, columns=df.columns)
    if '_is_bucket' not in df.columns:
        df = df.copy(); df['_is_bucket'] = False
        bucket_df['_is_bucket'] = True
    return pd.concat([df, bucket_df], ignore_index=True)


def run_pipeline_from_parsed(parsed, country, output_dir, cfg,
                             carrier_defaults=None, variables_layout=None,
                             exceptions=None, overflow_rules=None,
                             postcode_rules=None,
                             pallet_zones=None, pallet_defaults=None,
                             pallet_overrides=None, pallet_maut=None,
                             pallet_max_band_kg=None, express_only=False,
                             parcel_postcodes=None):
    """Build/optimize/write from an already-parsed rate dict.
    Used by the master-file path so the (expensive) parse happens only once.

    Parcel behaviour is unchanged. If `pallet_zones` is given (a
    {zip: {band_kg: rate}} dict from pallet_parser.country_pallet_data), DHL-FENDER
    pallet rows are built and merged into every stage. Whenever pallet rows are
    present the matrix is written NUMERICALLY (no Variables formulas) to match the
    reference pallet file and keep per-country surcharges exact.

    Always returns a 'minimal_df' key: the path to a pickled numeric DataFrame of
    the minimal matrix, which app.py concatenates into the combined workbook.
    """
    cd = carrier_defaults or CARRIER_DEFAULTS
    vl = variables_layout or VARIABLES_LAYOUT
    country = country.upper()

    df = build_extended_matrix(parsed, cfg)
    log.info('raw parcel rows: %d', len(df))
    if not df.empty:
        df = compute_numeric_totals(df, cd)

    # ── Express-only mode ─────────────────────────────────────────────────────
    # Keep only the EXPRESS SAVER service rows: UPDE 'EXPRESS SAVER 7R9W62' and
    # 'EXPRESS SAVER', UPSNL 'EXPRESS SAVER', UPSGB 'EXPRESS SAVER'. Drops
    # STANDARD / PARCEL / WORLDEASE and switches pallets off (express is
    # parcel-only). The substring match catches both UPDE express variants.
    if express_only:
        pallet_zones = None
        if not df.empty:
            keep = df['SERVICE_LEVEL'].astype(str).str.contains(
                'EXPRESS SAVER', case=False, na=False)
            df = df[keep].copy().reset_index(drop=True)
            log.info('express-only filter: %d express rows kept', len(df))

    # ── Pallet rows (optional) ───────────────────────────────────────────────
    df_pal_ext = pd.DataFrame()
    df_pal_opt = pd.DataFrame()
    pallet_warn = None
    if pallet_zones:
        try:
            import pallet_parser as _pp
            bands = sorted({int(b) for zmap in pallet_zones.values() for b in zmap})
        except Exception:
            bands = sorted({int(b) for zmap in pallet_zones.values() for b in zmap})
        if pallet_max_band_kg:                       # configurable max pallet weight
            bands = [b for b in bands if b <= pallet_max_band_kg]
        df_pal_ext, maut_known = build_pallet_df(
            country, pallet_zones, bands,
            pallet_defaults, pallet_overrides, pallet_maut)
        df_pal_opt = collapse_pallet_bands(df_pal_ext)
        if not maut_known and not df_pal_ext.empty:
            pallet_warn = (f"⚠️ DHL-FENDER: MAUT % unknown for {country} — pallet "
                           f"MAUT set to 0. Add it in the pallet MAUT table.")
        log.info('pallet rows: %d ext / %d collapsed', len(df_pal_ext), len(df_pal_opt))

    if df.empty and df_pal_ext.empty:
        raise ValueError(f"No rows built for {country} — no matching parcel or "
                         f"pallet rate data.")

    out = Path(output_dir)
    ext_path = out / f'{country}_Matrix_extended.xlsx'
    opt_path = out / f'{country}_Matrix_optimized.xlsx'
    min_path = out / f'{country}_Matrix_minimal.xlsx'

    df_opt = optimize_matrix(df) if not df.empty else pd.DataFrame()
    has_pallet = not df_pal_ext.empty

    any_buckets = bool(exceptions or overflow_rules or postcode_rules)
    if any_buckets and not df.empty:
        df_min = optimize_globally_df(df_opt)

        def _decorate(d):
            d = add_overflow_buckets(d, overflow_rules, cd, cfg)
            d = add_postcode_catchall(d, postcode_rules, cd)
            d = apply_exceptions(d, exceptions)
            return d

        df_ext_final = _decorate(df)
        df_opt_final = _decorate(df_opt)
        df_min_final = _decorate(df_min)
    else:
        df_min = optimize_globally_df(df_opt) if not df_opt.empty else pd.DataFrame()
        df_ext_final, df_opt_final, df_min_final = df, df_opt, df_min

    add_buckets = not express_only   # express-only builds are parcel-only

    # Parcel postcode expansion is applied to the MINIMAL stage only — that is the
    # deliverable CargoWrite consumes (and the source of the combined export).
    # Extended / optimized stay compact diagnostics. Pallet rows are untouched.
    pc_warn = None

    def _explode_min(frame):
        nonlocal pc_warn
        exploded, n = explode_parcel_postcodes(frame, parcel_postcodes)
        if n == -1 and pc_warn is None:
            pc_warn = (f"⚠️ {country}: no postcode list available — parcel rows "
                       f"left blank, so CargoWrite will skip them. Add {country} "
                       f"postcodes to the pallet/postcode file.")
        elif n > 0:
            log.info('%s: exploded %d blank parcel rows × %d postcodes',
                     country, n, len([p for p in (parcel_postcodes or []) if str(p).strip()]))
        return exploded

    if has_pallet:
        # Merge pallet rows into each stage (pallet rows never dominate parcel
        # rows and vice-versa — different shipment profiles), write with formulas.
        ext_all = _align_columns([df_ext_final, df_pal_ext])
        opt_all = _align_columns([df_opt_final, df_pal_opt])
        min_all = _align_columns([df_min_final, df_pal_opt])
        if add_buckets:
            ext_all = append_euroconnect_buckets(ext_all)
            opt_all = append_euroconnect_buckets(opt_all)
            min_all = append_euroconnect_buckets(min_all)
        min_all = _explode_min(min_all)
        write_matrix_with_formulas(ext_all, ext_path, cfg, cd, vl, pallet_maut,
                                   pallet_defaults)
        write_matrix_with_formulas(opt_all, opt_path, cfg, cd, vl, pallet_maut,
                                   pallet_defaults)
        write_matrix_with_formulas(min_all, min_path, cfg, cd, vl, pallet_maut,
                                   pallet_defaults)
        rows_ext, rows_opt, rows_min = len(ext_all), len(opt_all), len(min_all)
        extended_frame, optimized_frame, minimal_frame = ext_all, opt_all, min_all
    elif any_buckets:
        if add_buckets:
            df_ext_final = append_euroconnect_buckets(df_ext_final)
            df_opt_final = append_euroconnect_buckets(df_opt_final)
            df_min_final = append_euroconnect_buckets(df_min_final)
        df_min_final = _explode_min(df_min_final)
        write_matrix_excel(df_ext_final, ext_path, cfg, cd, vl)
        write_matrix_excel(df_opt_final, opt_path, cfg, cd, vl)
        write_matrix_excel(df_min_final, min_path, cfg, cd, vl)
        rows_ext, rows_opt, rows_min = (len(df_ext_final), len(df_opt_final),
                                        len(df_min_final))
        extended_frame, optimized_frame = df_ext_final, df_opt_final
        minimal_frame = df_min_final
    else:
        if add_buckets:
            df     = append_euroconnect_buckets(df)
            df_opt = append_euroconnect_buckets(df_opt)
            df_min = append_euroconnect_buckets(df_min)
        write_matrix_excel(df, ext_path, cfg, cd, vl)
        write_matrix_excel(df_opt, opt_path, cfg, cd, vl)
        # Use the in-memory minimal frame (optimize_globally_df, computed above)
        # rather than the optimize_globally Excel round-trip + read_excel: the
        # round-trip coerced POSTCODE to float and dropped leading zeros. Explode
        # parcel postcodes, then write numerically.
        df_min = _explode_min(df_min)
        write_matrix_excel(df_min, min_path, cfg, cd, vl)
        rows_ext, rows_opt, rows_min = len(df), len(df_opt), len(df_min)
        minimal_frame = df_min
        extended_frame, optimized_frame = df, df_opt

    # Persist the numeric stage frames for the combined-workbook export.
    def _persist_frame(frame, name):
        path = out / f'{country}_Matrix_{name}.pkl'
        try:
            frame.to_pickle(path)
            return str(path)
        except Exception:
            return None

    ext_df_str = _persist_frame(extended_frame,  'extended')
    opt_df_str = _persist_frame(optimized_frame, 'optimized')
    min_df_str = _persist_frame(minimal_frame,   'minimal')

    log.info('=== Done %s: %d ext / %d opt / %d min ===',
             country, rows_ext, rows_opt, rows_min)
    return {
        'extended':       str(ext_path),
        'optimized':      str(opt_path),
        'minimal':        str(min_path),
        'extended_df':    ext_df_str,
        'optimized_df':   opt_df_str,
        'minimal_df':     min_df_str,
        'rows_extended':  rows_ext,
        'rows_optimized': rows_opt,
        'rows_minimal':   rows_min,
        'pallet_warning': pallet_warn,
        'postcode_warning': pc_warn,
    }


# ==============================================================================
# 11. PALLET / FREIGHT INTEGRATION  (carrier DHL-FENDER, service EUROCONNECT)
# ==============================================================================
#
# Pallet rates come from the "DHL pricing with factor for DSV matrix" file via
# pallet_parser.country_pallet_data() -> {zip_prefix(str): {band_ceiling_kg: rate}}.
# These rates are already FACTORED. The cost stack on top (verified to the cent
# against ITFRDE_final_for_Fender_Pallets.xlsx):
#
#   RATE_BASE = factored_rate / FACTOR_DHL
#   FUEL      = fuel_pct      * RATE_BASE        (global)
#   MOBILITY  = mobility_pct  * RATE_BASE        (global)
#   MAUT      = maut_pct(country, band) * RATE_BASE   (per-country, 2-tier)
#   TOLL      = toll_pct(country) * RATE_BASE     (GB only in this contract)
#   ADMIN     = admin_per_shipment (€, global flat per row)
#   TOTAL     = RATE_BASE + FUEL + MOBILITY + MAUT + TOLL + ADMIN
#
# Pallet rows are NUMERIC (no Variables-sheet formulas) — matching the reference
# file and the combined export. When a matrix contains pallet rows the WHOLE
# matrix is written numerically; pure-parcel matrices keep the formula writer.

PALLET_DEFAULTS = {
    'DHL-FENDER': {
        'label':              'DHL Freight',
        'service_level':      'EUROCONNECT',
        'fuel_pct':           0.155,   # FUEL DHL PALLET
        'mobility_pct':       0.04,    # MOBILITY PALLET
        'admin_per_shipment': 46.51,   # ADMIN PALLET (€, applies to every row)
        'factor':             4.1278,     # FACTOR DHL (rates already factored)
        'toll_pct':           0.0,     # default no toll; GB set in overrides
    },
}

# Per-country surcharge overrides. Only the UK carries the road toll in this
# contract. app.py overwrites toll_pct / admin_per_shipment here from the sidebar.
PALLET_COUNTRY_OVERRIDES = {
    'GB': {'toll_pct': 0.0043, 'admin_per_shipment': 46.51},
}

# Per-country MAUT as a % of RATE_BASE, with an optional 2nd tier above a weight
# breakpoint:  (low_pct, high_pct, tier_kg)  — high_pct applies when band ceiling
# > tier_kg. Source: DSV pallet MAUT list (S2026). Low tier is 2.53% everywhere;
# the high tier (>2500 kg, up to FTL) is 5.44% for the road-toll countries below
# and 2.53% (flat) for the rest. Countries absent here default to 0 MAUT and raise
# a warning (never guessed).
PALLET_MAUT = {iso: (0.0253, 0.0544, 2500) for iso in
               ('AT', 'CH', 'CZ', 'DE', 'HR', 'HU', 'PL', 'SI', 'SK')}
PALLET_MAUT.update({iso: (0.0253, 0.0253, 2500) for iso in
                    ('BA', 'BG', 'DK', 'EE', 'FI', 'GR', 'IT', 'LT', 'LU', 'LV',
                     'MK', 'NO', 'RO', 'RS', 'SE', 'TR')})
# Confirmed by Fender logistics: these countries are genuinely 0% MAUT.
# Listed explicitly so they price correctly and stop firing "MAUT unknown" warnings.
PALLET_MAUT.update({iso: (0.0, 0.0, 2500) for iso in
                    ('BE', 'ES', 'FR', 'GB', 'IE', 'MT', 'NL', 'PT')})

# Extra columns pallet rows carry, in the reference file's order.
PALLET_COLUMN_ORDER = [
    'SITE_ID', 'CLIENT_ID', 'CARRIER_ID', 'SERVICE_LEVEL', 'COUNTRYISO2',
    'POSTCODE', 'MIN_WEIGHT', 'MAX_WEIGHT', 'MIN_VOLUME', 'MAX_VOLUME',
    'MIN_PARCEL', 'MAX_PARCEL', 'EACH_WEIGHT', 'EACH_VOLUME',
    'FACTORED RATE PALLET', 'USER_DEF_TYPE_1', 'USER_DEF_TYPE_2',
    'USER_DEF_TYPE_4 (max 1,5m)', 'AWKWARD',
    'RATE_BASE', 'RATE_EXTRA', 'MOBILITY', 'FUEL', 'MAUT', 'Linehaul UPSDE',
    'TOLL', 'ADMIN', 'TOTAL_PRICE',
]


def _pallet_maut_for(country, ceiling_kg, maut_table):
    rule = maut_table.get(country.upper())
    if rule is None:
        return None                       # unknown -> caller warns
    low, high, tier = rule
    return high if (ceiling_kg is not None and ceiling_kg > tier) else low


def build_pallet_df(country, zip_rate_map, band_ceilings,
                    pallet_defaults=None, pallet_overrides=None,
                    pallet_maut=None):
    """Return (DataFrame, maut_known) of DHL-FENDER pallet rows for one country.

    zip_rate_map : {zip_prefix(str): {ceiling_kg(int): factored_rate}}
    band_ceilings: ordered list of band ceilings (kg)
    """
    pd_def = (pallet_defaults or PALLET_DEFAULTS)['DHL-FENDER']
    ov     = (pallet_overrides or PALLET_COUNTRY_OVERRIDES).get(country.upper(), {})
    mt     = pallet_maut or PALLET_MAUT
    iso    = country.upper()

    fuel_pct  = pd_def['fuel_pct']
    mob_pct   = pd_def['mobility_pct']
    factor    = pd_def['factor']
    toll_pct  = ov.get('toll_pct', pd_def.get('toll_pct', 0.0))
    admin     = ov.get('admin_per_shipment', pd_def['admin_per_shipment'])
    service   = pd_def['service_level']
    maut_known = iso in mt

    rows = []
    for zkey in sorted(zip_rate_map, key=lambda z: (len(str(z)), str(z))):
        band_map = zip_rate_map[zkey]
        prev_ceiling = 0
        for ceil in band_ceilings:
            rate = band_map.get(ceil)
            if rate is None:
                continue
            rate_base = round(rate / factor, 6)
            fuel = round(fuel_pct * rate_base, 6)
            mob  = round(mob_pct * rate_base, 6)
            maut_pct = _pallet_maut_for(iso, ceil, mt) or 0.0
            maut = round(maut_pct * rate_base, 8)
            toll = round(toll_pct * rate_base, 6)
            total = round(rate_base + fuel + mob + maut + toll + admin, 6)
            rows.append({
                'SITE_ID': 'NLMOE01', 'CLIENT_ID': 'NLFENDER',
                'CARRIER_ID': 'DHL-FENDER', 'SERVICE_LEVEL': service,
                'COUNTRYISO2': iso, 'POSTCODE': str(zkey),
                'MIN_WEIGHT': prev_ceiling if prev_ceiling > 0 else None,
                'MAX_WEIGHT': ceil, 'MIN_VOLUME': None, 'MAX_VOLUME': None,
                'MIN_PARCEL': None, 'MAX_PARCEL': None,
                'EACH_WEIGHT': None, 'EACH_VOLUME': None,
                'FACTORED RATE PALLET': rate_base,
                'USER_DEF_TYPE_1': None, 'USER_DEF_TYPE_2': None,
                'USER_DEF_TYPE_4 (max 1,5m)': None, 'AWKWARD': None,
                'RATE_BASE': rate_base, 'RATE_EXTRA': 0,
                'MOBILITY': mob, 'FUEL': fuel, 'MAUT': maut,
                'Linehaul UPSDE': None, 'TOLL': toll, 'ADMIN': admin,
                'TOTAL_PRICE': total, '_is_bucket': False,
            })
            prev_ceiling = ceil
    return pd.DataFrame(rows), maut_known


def collapse_pallet_bands(df_pallet):
    """Within each (country, zip), drop a band whose RATE_BASE equals the band
    directly below it — CargoWrite matches the first row with MAX_WEIGHT >= wt,
    so the lighter identical-priced band already covers those shipments."""
    if df_pallet.empty:
        return df_pallet
    df = df_pallet.sort_values(['COUNTRYISO2', 'POSTCODE', 'MAX_WEIGHT'],
                               kind='stable').reset_index(drop=True)
    keep, last_key, last_rate = [], None, None
    for i, r in df.iterrows():
        key = (r['COUNTRYISO2'], r['POSTCODE'])
        if key == last_key and r['RATE_BASE'] == last_rate:
            continue
        keep.append(i)
        last_key, last_rate = key, r['RATE_BASE']
    return df.loc[keep].reset_index(drop=True)


def _align_columns(frames):
    """Concat frames giving them a common column set (union), pallet order if any
    pallet column is present, else the parcel COLUMN_ORDER."""
    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame()
    has_pallet = any('FACTORED RATE PALLET' in f.columns for f in frames)
    order = PALLET_COLUMN_ORDER if has_pallet else COLUMN_ORDER
    cols = list(order) + (['_is_bucket'] if any('_is_bucket' in f.columns for f in frames) else [])
    out = []
    for f in frames:
        g = f.copy()
        for c in cols:
            if c not in g.columns:
                g[c] = None
        out.append(g[cols])
    return pd.concat(out, ignore_index=True)


def write_matrix_numeric(df, output_path, country_cfg, variables_layout=None,
                         column_order=None):
    """Write a matrix as NUMERIC values (no formulas). Used whenever pallet rows
    are present, and for the combined export. Colours bucket rows amber."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    vl    = variables_layout or VARIABLES_LAYOUT
    order = column_order or (PALLET_COLUMN_ORDER
                             if 'FACTORED RATE PALLET' in df.columns else COLUMN_ORDER)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{country_cfg.get('iso2', 'ALL')} Matrix"
    for ci, col in enumerate(order, 1):
        ws.cell(1, ci, col)
    df_sorted = df.sort_values('TOTAL_PRICE', kind='stable').reset_index(drop=True)
    fill = PatternFill('solid', fgColor='FFF2CC')
    for ri, rec in enumerate(df_sorted.to_dict('records'), start=2):
        is_bucket = bool(rec.get('_is_bucket'))
        for ci, col in enumerate(order, 1):
            v = rec.get(col)
            cell = ws.cell(ri, ci, None if (v is None or (isinstance(v, float) and pd.isna(v))) else v)
            if is_bucket:
                cell.fill = fill
    vs = wb.create_sheet('Variables')
    for ri, (name, val) in enumerate(vl, 1):
        vs.cell(ri, 1, name)
        if val is not None:
            vs.cell(ri, 2, val)
    wb.save(output_path)
    log.info('wrote %s (%d rows, numeric)', output_path, len(df_sorted))


def write_combined_matrix(frames, output_path, variables_layout=None,
                          pallet_maut=None, pallet_defaults=None,
                          carrier_defaults=None, formulas=True):
    """Merge every country's minimal frame into ONE sheet, sorted by country then
    price. With formulas=True (default) the sheet uses live Variables formulas;
    per-country pallet MAUT cells are written into the Variables sheet. `frames`
    is a list of DataFrames (from result['minimal_df'])."""
    combined = _align_columns(frames)
    if combined.empty:
        raise ValueError("write_combined_matrix: no rows to write.")
    has_pallet = 'FACTORED RATE PALLET' in combined.columns
    order = PALLET_COLUMN_ORDER if has_pallet else COLUMN_ORDER
    combined = combined.sort_values(['COUNTRYISO2', 'TOTAL_PRICE'],
                                    kind='stable').reset_index(drop=True)
    if formulas and has_pallet:
        write_matrix_with_formulas(combined, output_path, {'iso2': 'ALL'},
                                   carrier_defaults, variables_layout,
                                   pallet_maut, pallet_defaults, column_order=order)
    else:
        write_matrix_numeric(combined, output_path, {'iso2': 'ALL'},
                             variables_layout, column_order=order)
    return str(output_path)


# ==============================================================================
# 12. FORMULA WRITER FOR PALLET-INCLUSIVE MATRICES
# ==============================================================================
#
# Pure-parcel matrices already use write_matrix_excel (formula writer, 22-col
# layout). When pallet rows are present the layout is 28 columns, so column
# letters shift and the parcel formula writer can't be reused as-is. This writer
# handles BOTH row types in the 28-col layout and references the Variables sheet:
#
#   FUEL/MOBILITY/TOLL  = <Variables pct cell> * RATE_BASE
#   MAUT (pallet)       = per-country, two-tier: references that country's MAUT
#                         low/high cell depending on the row's MAX_WEIGHT vs tier
#   MAUT/FUEL (parcel)  = the carrier's existing Variables refs (B1..B9)
#   ADMIN               = <Variables ADMIN cell>  (flat constant)
#   volumes             = weight / divisor   (parcel rows only)
#   TOTAL               = sum of the components present for that row type
#
# Editing any Variables cell recomputes the whole sheet in Excel.

def _letter_map(column_order):
    import openpyxl.utils as _u
    return {name: _u.get_column_letter(i + 1) for i, name in enumerate(column_order)}


def _ensure_var(vars_rows, name, default):
    """Return (vars_rows, row_index_1based) ensuring `name` exists in the layout."""
    for i, (n, _v) in enumerate(vars_rows):
        if n == name:
            return vars_rows, i + 1
    vars_rows = list(vars_rows) + [(name, default)]
    return vars_rows, len(vars_rows)


def write_matrix_with_formulas(df, output_path, country_cfg,
                               carrier_defaults=None, variables_layout=None,
                               pallet_maut=None, pallet_defaults=None,
                               column_order=None):
    """Write a pallet-inclusive matrix with live formulas referencing Variables."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    cd    = carrier_defaults or CARRIER_DEFAULTS
    pdef  = (pallet_defaults or PALLET_DEFAULTS)['DHL-FENDER']
    mt    = pallet_maut or PALLET_MAUT
    order = column_order or PALLET_COLUMN_ORDER
    L     = _letter_map(order)

    vars_rows = list(variables_layout or VARIABLES_LAYOUT)
    # Ensure the pallet global cells exist and capture their Variables rows.
    vars_rows, r_fuel  = _ensure_var(vars_rows, 'FUEL DHL PALLET', pdef['fuel_pct'])
    vars_rows, r_mob   = _ensure_var(vars_rows, 'MOBILITY PALLET', pdef['mobility_pct'])
    vars_rows, r_toll  = _ensure_var(vars_rows, 'TOLL UK PALLET', pdef.get('toll_pct', 0.0043))
    vars_rows, r_admin = _ensure_var(vars_rows, 'ADMIN PALLET', pdef['admin_per_shipment'])

    # Per-country pallet MAUT block: name | low(B) | high(C) | tier(D)
    # Compute the Variables rows up-front so row formulas can reference them.
    pallet_countries = sorted(
        set(df.loc[df['CARRIER_ID'] == 'DHL-FENDER', 'COUNTRYISO2'].dropna())
    ) if 'CARRIER_ID' in df.columns else []
    maut_row = {}        # iso -> Variables row (1-based)
    maut_header_row = None
    if pallet_countries:
        maut_header_row = len(vars_rows) + 2         # one spacer row after globals
        cur = maut_header_row + 1
        for iso in pallet_countries:
            maut_row[iso] = cur
            cur += 1

    # write Variables sheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{country_cfg.get('iso2', 'ALL')} Matrix"
    for ci, col in enumerate(order, 1):
        ws.cell(1, ci, col)

    df_sorted = df.sort_values('TOTAL_PRICE', kind='stable').reset_index(drop=True)
    fill = PatternFill('solid', fgColor='FFF2CC')

    L_RATE  = L['RATE_BASE']
    L_EXTRA = L.get('RATE_EXTRA')
    L_FUEL  = L['FUEL']
    L_MAUT  = L['MAUT']
    L_MOB   = L.get('MOBILITY')
    L_TOLL  = L.get('TOLL')
    L_ADMIN = L.get('ADMIN')
    L_LH    = L.get('Linehaul UPSDE')
    L_MP    = L.get('MAX_PARCEL')
    L_EW    = L.get('EACH_WEIGHT')
    L_MW    = L.get('MAX_WEIGHT')
    L_MV    = L.get('MAX_VOLUME')
    L_EV    = L.get('EACH_VOLUME')

    for ri, rec in enumerate(df_sorted.to_dict('records'), start=2):
        carrier   = rec.get('CARRIER_ID')
        is_bucket = bool(rec.get('_is_bucket'))
        is_pallet = (carrier == 'DHL-FENDER')
        # Sentinel catch-all bucket (no rate components): keep its literal price.
        sentinel = is_bucket and pd.isna(rec.get('RATE_BASE'))
        formulas = {}

        if sentinel:
            pass
        elif is_pallet:
            iso = rec.get('COUNTRYISO2')
            formulas['FUEL']     = f"=Variables!$B${r_fuel}*{L_RATE}{ri}"
            formulas['MOBILITY'] = f"=Variables!$B${r_mob}*{L_RATE}{ri}"
            # two-tier MAUT: pick low/high by this row's band vs the country tier
            low, high, tier = mt.get(iso, (0.0, 0.0, 2500))
            mw = rec.get('MAX_WEIGHT')
            r_iso = maut_row.get(iso)
            col = 'C' if (mw is not None and not pd.isna(mw) and mw > tier) else 'B'
            formulas['MAUT'] = f"=Variables!${col}${r_iso}*{L_RATE}{ri}"
            # TOLL: GB carries it; others 0 (still summed)
            if L_TOLL:
                if (rec.get('TOLL') or 0) > 0:
                    formulas['TOLL'] = f"=Variables!$B${r_toll}*{L_RATE}{ri}"
            if L_ADMIN:
                formulas['ADMIN'] = f"=Variables!$B${r_admin}"
            parts = [f"{L_RATE}{ri}"]
            if L_EXTRA: parts.append(f"{L_EXTRA}{ri}")
            if L_MOB:   parts.append(f"{L_MOB}{ri}")
            parts.append(f"{L_FUEL}{ri}")
            parts.append(f"{L_MAUT}{ri}")
            if L_TOLL:  parts.append(f"{L_TOLL}{ri}")
            if L_ADMIN: parts.append(f"{L_ADMIN}{ri}")
            formulas['TOTAL_PRICE'] = "=" + "+".join(parts)
        else:
            cfg = cd.get(carrier, {})
            has_grid = (rec.get('MAX_PARCEL') is not None and not pd.isna(rec.get('MAX_PARCEL'))
                        and rec.get('EACH_WEIGHT') is not None and not pd.isna(rec.get('EACH_WEIGHT')))
            if has_grid and L_MW and L_MP and L_EW:
                formulas['MAX_WEIGHT'] = f"={L_MP}{ri}*{L_EW}{ri}"
                if L_MV: formulas['MAX_VOLUME']  = f"={L_MW}{ri}/{cfg.get('volume_divisor', 1)}"
                if L_EV: formulas['EACH_VOLUME'] = f"={L_EW}{ri}/{cfg.get('volume_divisor', 1)}"
            if cfg.get('fuel_variables_ref'):
                ref = cfg['fuel_variables_ref']
                formulas['FUEL'] = f"=Variables!${ref[0]}${ref[1:]}*{L_RATE}{ri}"
            if cfg.get('maut_variables_ref'):
                ref = cfg['maut_variables_ref']
                formulas['MAUT'] = f"=Variables!${ref[0]}${ref[1:]}*{L_RATE}{ri}"
            parts = [f"{L_RATE}{ri}"]
            if L_EXTRA: parts.append(f"{L_EXTRA}{ri}")
            parts.append(f"{L_FUEL}{ri}")
            parts.append(f"{L_MAUT}{ri}")
            if L_LH: parts.append(f"{L_LH}{ri}")
            formulas['TOTAL_PRICE'] = "=" + "+".join(parts)

        for ci, col in enumerate(order, 1):
            if col in formulas:
                cell = ws.cell(ri, ci, formulas[col])
            else:
                v = rec.get(col)
                cell = ws.cell(ri, ci, None if (v is None or (isinstance(v, float) and pd.isna(v))) else v)
            if is_bucket:
                cell.fill = fill

    # ── Variables sheet ──────────────────────────────────────────────────────
    vs = wb.create_sheet('Variables')
    for r_i, (name, val) in enumerate(vars_rows, 1):
        vs.cell(r_i, 1, name)
        if val is not None:
            vs.cell(r_i, 2, val)
    # MAUT block with low(B)/high(C)/tier(D) at the rows reserved earlier
    if pallet_countries:
        vs.cell(maut_header_row, 1, 'MAUT DHL PALLET — low(B) / high(C) / tier kg(D)')
        for iso in pallet_countries:
            low, high, tier = mt.get(iso, (0.0, 0.0, 2500))
            r = maut_row[iso]
            vs.cell(r, 1, f'MAUT DHL PALLET {iso}')
            vs.cell(r, 2, low)
            vs.cell(r, 3, high)
            vs.cell(r, 4, tier)

    wb.save(output_path)
    log.info('wrote %s (%d rows, formulas)', output_path, len(df_sorted))
