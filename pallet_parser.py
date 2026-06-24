"""
pallet_parser.py — parse the "DHL pricing with factor for DSV matrix" workbook.

This is the SINGLE source of truth for DHL freight (pallet / EUROCONNECT) rates.
Unlike the parcel cards, every country and postcode prefix already carries a
fully *factored* rate per weight band — so there is no zone lookup to do: the
(country, zip-prefix) -> {band_ceiling_kg: rate} mapping is read directly.

Reconciliation note (verified against ITFRDE_final_for_Fender_Pallets.xlsx):
    DE / zip 52 / band <=100 kg  ->  70.931066   (matches to the cent)

Public API
----------
is_pallet_factor_file(path)      -> bool
parse_pallet_factor_file(path)   -> dict   (parsed once)
available_pallet_countries(p)    -> set
country_pallet_data(parsed, iso) -> dict   {zip_prefix(str): {ceil_kg(int): rate}}
band_ceilings(parsed)            -> list[int]   ordered band ceilings
"""

import re
import logging

import openpyxl

log = logging.getLogger(__name__)

_BAND_RE = re.compile(r'-\s*([\d.,]+)\s*kg', re.IGNORECASE)
_ISO2 = re.compile(r'[A-Z]{2}')


def is_pallet_factor_file(path):
    """True if this is the DHL factor workbook (Country / Zip / '<n> kg' columns)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    ws = wb[wb.sheetnames[0]]
    hdr = next(ws.iter_rows(max_row=1, values_only=True), ())
    hdr_norm = [str(h).strip().lower() if h is not None else '' for h in hdr]
    has_country = 'country' in hdr_norm
    has_zip = 'zip' in hdr_norm or 'postcode' in hdr_norm
    has_band = any('kg' in h for h in hdr_norm)
    return has_country and has_zip and has_band


def _band_ceiling(header_value):
    """'100,1 - 200 kg' -> 200 ;  'FTL' -> None."""
    if header_value is None:
        return None
    m = _BAND_RE.search(str(header_value))
    if not m:
        return None
    raw = m.group(1).replace('.', '').replace(',', '')
    try:
        return int(raw)
    except ValueError:
        return None


def parse_pallet_factor_file(path):
    """Parse the whole workbook once.

    Returns
    -------
    {
      'bands':   [100, 200, 300, ...],          # ordered ceilings (kg)
      'has_ftl': bool,
      'rates': { 'DE': { '52': {100: 70.931066, 200: ..., ...}, ... }, ... },
    }
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)

    # locate Country / Zip columns
    col_country = col_zip = None
    band_cols = {}          # col_index -> ceiling_kg
    has_ftl = False
    for c, h in enumerate(header):
        hn = str(h).strip().lower() if h is not None else ''
        if hn == 'country':
            col_country = c
        elif hn in ('zip', 'postcode'):
            col_zip = c
        elif hn == 'ftl':
            has_ftl = True
        else:
            ceil = _band_ceiling(h)
            if ceil is not None:
                band_cols[c] = ceil

    if col_country is None or col_zip is None or not band_cols:
        raise ValueError("Not a recognisable DHL factor file (missing Country/Zip/bands).")

    bands = sorted(set(band_cols.values()))
    rates = {}
    for row in rows:
        country = row[col_country]
        if not (isinstance(country, str) and _ISO2.fullmatch(country.strip())):
            continue
        iso = country.strip().upper()
        zraw = row[col_zip]
        if zraw is None:
            continue
        # zip prefixes are stored as text ('01','52') — keep leading zeros
        zkey = str(zraw).strip()
        if zkey == '':
            continue
        band_map = {}
        for c, ceil in band_cols.items():
            v = row[c]
            if isinstance(v, (int, float)):
                band_map[ceil] = float(v)
        if band_map:
            rates.setdefault(iso, {})[zkey] = band_map

    log.info("pallet factor file: %d countries, %d bands%s",
             len(rates), len(bands), " (+FTL)" if has_ftl else "")
    return {'bands': bands, 'has_ftl': has_ftl, 'rates': rates}


def available_pallet_countries(parsed):
    return set(parsed.get('rates', {}).keys())


def country_pallet_data(parsed, iso2):
    return parsed.get('rates', {}).get(iso2.upper(), {})


def band_ceilings(parsed):
    return list(parsed.get('bands', []))


# ── Backward-compat alias ─────────────────────────────────────────────────────
# app.py calls pp.parse_pallet_rate_card(); keep one canonical implementation.
parse_pallet_rate_card = parse_pallet_factor_file
is_pallet_rate_card    = is_pallet_factor_file
